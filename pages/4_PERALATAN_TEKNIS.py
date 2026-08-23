import os
import io
import re
import tempfile
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- KONFIGURASI HALAMAN STREAMLIT ---
st.set_page_config(
    page_title="Pengolah Data ALOPTAMA",
    page_icon="🌤️",
    layout="wide"
)

# --- ENGINE OLAH DATA & HELPER UNIVERSAL ---
class EnginePerapihData:
    
    KOLOM_AWS_RESMI = [
        'Date', 'Time', 'S', 'DD', 'FF', 'DM10', 'FM10', 'DD2', 'FF2', 
        'DVN', 'DVX', 'FVN', 'FVX', 'RR', 'RH', 'TSV', 'DP', 'T', 
        'GLOR', 'GLORP', 'INSD', 'STAP', 'MSLP', 'GNDT'
    ]

    UNIVERSAL_LABELS = {
        'Air_Temp': 'Suhu Udara',
        'Dew_Point': 'Titik Embun',
        'RH': 'Kelembaban Udara',
        'QFE': 'Tekanan QFE',
        'QNH': 'Tekanan QNH',
        'Wind_Speed': 'Kecepatan Angin',
        'Wind_Dir': 'Arah Angin',
        'Solar_Rad': 'Radiasi Matahari',
        'Precip_1Hr': 'Curah Hujan',
        'Ground_Temp': 'Suhu Tanah'
    }

    @staticmethod
    def standardize_columns(df):
        """Memetakan sensor utama ke nama universal & membersihkan suffix runway (33, 07, .RWYA) dari semua kolom."""
        rename_dict = {}
        
        for c in df.columns:
            str_c = str(c).strip()
            
            # Abaikan penamaan ulang universal jika kolom merupakan sub-metric agregasi (10 Min, 2 Min, MAX, MIN, 24 Hr)
            is_submetric = re.search(r'10\s*min|2\s*min|60\s*min|1\s*min|24\s*hr|max|min|avg|gust', str_c, re.I)
            
            if not is_submetric and re.search(r'^(air\s*tm?p|airtemperaturevalidated|T)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'Air_Temp'
            elif not is_submetric and re.search(r'^(dew\s*pt|dewpointvalidated|DP)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'Dew_Point'
            elif not is_submetric and re.search(r'^(relativehumidityvalidated|relative\s*humidity|RH)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'RH'
            elif not is_submetric and re.search(r'^(windspeedselected|wind\s*speed|WS|FF)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'Wind_Speed'
            elif not is_submetric and re.search(r'^(instant\.winddirection|wind\s*direction|mag\s*wd|WD|DD)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'Wind_Dir'
            elif not is_submetric and re.search(r'^(QFE|STAP)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'QFE'
            elif not is_submetric and re.search(r'^(QNH)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'QNH'
            elif not is_submetric and re.search(r'^(solarradiation|solar\s*rad|GLOR)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'Solar_Rad'
            elif not is_submetric and re.search(r'^(onehour\.precipselected|precip\s*1hr|RR)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'Precip_1Hr'
            elif not is_submetric and re.search(r'^(ground\s*temp|GNDT)(\s*\(.*\))?(\s*\d{1,2}|\.RWY[A-Z0-9]+)?$', str_c, re.I):
                rename_dict[c] = 'Ground_Temp'
            else:
                # Membersihkan suffix runway (seperti " 33", ".RWYA", " RWYA") dari seluruh kolom pendukung
                s = re.sub(r'(\.|\s+)RWY[A-Z0-9]*', '', str_c, flags=re.I)
                s = re.sub(r'(\)\s*)\d{1,2}\b', r'\1', s)
                s = re.sub(r'\s+\d{1,2}$', '', s)
                s = re.sub(r'\s+', ' ', s).strip()
                rename_dict[c] = s

        df_renamed = df.rename(columns=rename_dict)
        
        # Mencegah duplikasi nama kolom jika ada lebih dari satu sensor sejenis
        cols = []
        counts = {}
        for col in df_renamed.columns:
            if col in counts:
                counts[col] += 1
                cols.append(f"{col}_{counts[col]}")
            else:
                counts[col] = 0
                cols.append(col)
        df_renamed.columns = cols
        return df_renamed

    @staticmethod
    def smart_parse_datetime(date_series, time_series):
        combined = date_series.astype(str).str.strip() + ' ' + time_series.astype(str).str.strip()
        dt_parsed = pd.to_datetime(combined, format='mixed', dayfirst=True, errors='coerce')
        return dt_parsed.dt.tz_localize('UTC', ambiguous='NaT', nonexistent='NaT')

    @classmethod
    def load_files_to_dataframe(cls, uploaded_files):
        list_df = []
        for fp in uploaded_files:
            ext = fp.name.lower()
            if ext.endswith('.fdb'):
                df_fdb = cls.baca_database_fdb(fp.getbuffer(), cls.KOLOM_AWS_RESMI)
                if not df_fdb.empty:
                    list_df.append(df_fdb)
            elif ext.endswith(('.xlsx', '.xls')):
                df_ex = pd.read_excel(fp)
                if len(df_ex.columns) == 1:
                    lines = df_ex.iloc[:, 0].dropna().astype(str).tolist()
                    rows = cls.parse_text_lines(lines, cls.KOLOM_AWS_RESMI)
                    if rows: list_df.append(pd.DataFrame(rows, columns=cls.KOLOM_AWS_RESMI))
                else:
                    if 'Date and Time' in df_ex.columns:
                        df_ex.rename(columns={'Date and Time': 'Date_Time_Raw'}, inplace=True)
                    if 'Date_Time_Raw' in df_ex.columns:
                        split_dt = df_ex['Date_Time_Raw'].astype(str).str.split(r'\s+', n=1, expand=True)
                        df_ex.insert(0, 'Time', split_dt[1].fillna('00:00:00') if split_dt.shape[1] > 1 else '00:00:00')
                        df_ex.insert(0, 'Date', split_dt[0])
                        df_ex.drop(columns=['Date_Time_Raw'], inplace=True)
                    list_df.append(df_ex)
            elif ext.endswith('.csv'):
                df_csv = pd.read_csv(fp)
                if 'Date' not in df_csv.columns:
                    orig_col = df_csv.columns[0]
                    split_dt = df_csv[orig_col].astype(str).str.split(r'\s+', n=1, expand=True)
                    df_csv.insert(0, 'Time', split_dt[1].fillna('00:00:00') if split_dt.shape[1] > 1 else '00:00:00')
                    df_csv.insert(0, 'Date', split_dt[0])
                    df_csv.drop(columns=[orig_col], inplace=True)
                list_df.append(df_csv)
            else:
                lines = fp.getvalue().decode('utf-8', errors='ignore').splitlines()
                rows = cls.parse_text_lines(lines, cls.KOLOM_AWS_RESMI)
                if rows: list_df.append(pd.DataFrame(rows, columns=cls.KOLOM_AWS_RESMI))
                
        return pd.concat(list_df, ignore_index=True) if list_df else pd.DataFrame()

    @staticmethod
    @st.cache_data(show_spinner=False)
    def normalize_dataframe(df):
        if df is None or df.empty or 'Date' not in df.columns or 'Time' not in df.columns:
            return df

        date_str = df['Date'].astype(str).str.strip()
        valid_mask = df['Date'].notna() & date_str.ne('') & date_str.str.lower().ne('nan')
        df_clean = df[valid_mask].copy()

        # Normalisasi Nama Kolom Universal & Pembersihan Suffix Runway
        df_clean = EnginePerapihData.standardize_columns(df_clean)

        # UTC Datetime Processing
        df_clean['datetime_temp'] = EnginePerapihData.smart_parse_datetime(df_clean['Date'], df_clean['Time'])
        df_clean.dropna(subset=['datetime_temp'], inplace=True)
        df_clean.drop_duplicates(subset=['datetime_temp'], inplace=True)
        df_clean.set_index('datetime_temp', inplace=True)
        df_clean.sort_index(inplace=True)

        if not df_clean.empty:
            df_clean = df_clean.resample('1min').asfreq()
        
        df_clean.reset_index(inplace=True)
        df_clean['Date'] = df_clean['datetime_temp'].dt.strftime('%Y-%m-%d')
        df_clean['Time'] = df_clean['datetime_temp'].dt.strftime('%H:%M:00')

        # Keys Power BI (UTC Basis)
        df_clean['PK_Datetime'] = df_clean['datetime_temp'].dt.strftime('%Y%m%d%H%M').astype('int64')
        df_clean['DateKey'] = df_clean['datetime_temp'].dt.strftime('%Y%m%d').astype('int64')
        df_clean['TimeKey'] = df_clean['datetime_temp'].dt.strftime('%H%M').astype('int64')

        # Type Casting Numerik
        cols_to_exclude = {'PK_Datetime', 'DateKey', 'TimeKey', 'Date', 'Time', 'datetime_temp', 'S'}
        target_cols = [c for c in df_clean.columns if c not in cols_to_exclude]
        for col in target_cols:
            s = df_clean[col]
            if isinstance(s, pd.DataFrame):
                s = s.iloc[:, 0]
            df_clean[col] = pd.to_numeric(s, errors='coerce')

        # QC Limits Check Universal
        qc_limits = {
            'Air_Temp': (-50, 60),
            'RH': (0, 100),
            'Wind_Dir': (0, 360),
            'Wind_Speed': (0, 150),
            'QFE': (600, 1150),
            'QNH': (600, 1150),
            'Ground_Temp': (-10, 70)
        }
        for col, (vmin, vmax) in qc_limits.items():
            if col in df_clean.columns:
                val = df_clean[col]
                if isinstance(val, pd.DataFrame):
                    val = val.iloc[:, 0]
                df_clean.loc[(val < vmin) | (val > vmax), col] = np.nan

        # Dew Point Calculation
        if 'Air_Temp' in df_clean.columns and 'RH' in df_clean.columns:
            valid_trh = df_clean['Air_Temp'].notna() & df_clean['RH'].notna()
            a, b = 17.27, 237.7
            alpha = ((a * df_clean['Air_Temp']) / (b + df_clean['Air_Temp'])) + np.log(np.maximum(df_clean['RH'], 1e-5) / 100.0)
            dp_calc = ((b * alpha) / (a - alpha)).round(1)
            
            if 'Dew_Point' not in df_clean.columns:
                df_clean['Dew_Point'] = np.nan
                
            missing_dp = df_clean['Dew_Point'].isna()
            df_clean.loc[missing_dp & valid_trh, 'Dew_Point'] = dp_calc[missing_dp & valid_trh]

        df_clean.drop(columns=['datetime_temp'], inplace=True, errors='ignore')
        return df_clean

    @staticmethod
    def generate_dim_date(df_clean):
        if df_clean.empty or 'DateKey' not in df_clean.columns:
            return pd.DataFrame()
        
        dates = pd.to_datetime(df_clean['DateKey'].astype(str), format='%Y%m%d').unique()
        dim_date = pd.DataFrame({'Date': pd.to_datetime(dates)})
        dim_date['DateKey'] = dim_date['Date'].dt.strftime('%Y%m%d').astype('int64')
        dim_date['Year'] = dim_date['Date'].dt.year
        dim_date['MonthNumber'] = dim_date['Date'].dt.month
        dim_date['MonthName'] = dim_date['Date'].dt.strftime('%B')
        dim_date['DayOfMonth'] = dim_date['Date'].dt.day
        dim_date['DayOfWeek'] = dim_date['Date'].dt.strftime('%A')
        dim_date['Quarter'] = 'Q' + dim_date['Date'].dt.quarter.astype(str)
        dim_date['IsWeekend'] = dim_date['Date'].dt.dayofweek.isin([5, 6])
        dim_date['Timezone'] = 'UTC'
        
        cols = ['DateKey', 'Date', 'Year', 'MonthNumber', 'MonthName', 'DayOfMonth', 'DayOfWeek', 'Quarter', 'IsWeekend', 'Timezone']
        return dim_date[cols]

    @staticmethod
    def generate_dim_time():
        utc_range = pd.date_range('2000-01-01 00:00', '2000-01-01 23:59', freq='1min')
        wita_range = utc_range + pd.Timedelta(hours=8)
        
        dim_time = pd.DataFrame({
            'TimeKey': utc_range.strftime('%H%M').astype('int64'),
            'Time_UTC': utc_range.strftime('%H:%M:00'),
            'Time_WITA': wita_range.strftime('%H:%M:00'),
            'Hour_WITA': wita_range.hour,
        })
        
        dim_time['PeriodOfDay_WITA'] = pd.cut(
            dim_time['Hour_WITA'],
            bins=[-1, 5, 11, 17, 23],
            labels=['Dini Hari', 'Pagi', 'Siang/Sore', 'Malam']
        )
        return dim_time[['TimeKey', 'Time_UTC', 'Time_WITA', 'Hour_WITA', 'PeriodOfDay_WITA']]

    @staticmethod
    def generate_dim_sensor(sensor_labels):
        qc_limits = {
            'Air_Temp': (-50, 60), 'Dew_Point': (-50, 60),
            'RH': (0, 100), 'Wind_Dir': (0, 360), 'Wind_Speed': (0, 150),
            'QFE': (600, 1150), 'QNH': (600, 1150), 'Ground_Temp': (-10, 70)
        }
        rows = [
            {
                'SensorCode': code,
                'SensorName': name,
                'QC_Min_Limit': qc_limits.get(code, (None, None))[0],
                'QC_Max_Limit': qc_limits.get(code, (None, None))[1]
            }
            for code, name in sensor_labels.items()
        ]
        return pd.DataFrame(rows)

    @staticmethod
    def parse_text_lines(lines, kolom_aws_resmi):
        if not lines: 
            return []
        
        data_rows = []
        start_idx = 0
        first_line = lines[0].strip()
        tokens_first = first_line.split()
        
        if tokens_first:
            t0 = tokens_first[0].lower()
            if any(k in t0 for k in ['date', 'tgl', 'tanggal', 'waktu']) or not any(c.isdigit() for c in t0):
                start_idx = 1

        needed_payload_len = len(kolom_aws_resmi) - 2
        
        for baris in lines[start_idx:]:
            tokens = baris.strip().split()
            if not tokens: 
                continue
            
            if len(tokens) > 1 and ':' in tokens[1]:
                date_val, time_val = tokens[0], tokens[1]
                remaining = tokens[2:]
            else:
                date_val, time_val = tokens[0], '00:00:00'
                remaining = tokens[1:]
            
            pad_len = needed_payload_len - len(remaining)
            if pad_len > 0:
                remaining.extend([''] * pad_len)
                
            data_rows.append([date_val, time_val] + remaining[:needed_payload_len])
            
        return data_rows

    @classmethod
    @st.cache_data(show_spinner=False)
    def buat_rangkuman_per_sensor(cls, df, sensor_labels):
        TOTAL_HARUSNYA_PER_HARI = 1440
        valid_sensors = [s for s in sensor_labels.keys() if s in df.columns]
        
        valid_df = df[df['Date'].notna() & (df['Date'] != '')].copy()
        if valid_df.empty:
            cols = ['Date', 'Total Menit Log'] + [f'{sensor_labels[s]} (%)' for s in valid_sensors] + ['PERSENTASE TOTAL KESELURUHAN (%)']
            return pd.DataFrame(columns=cols)

        counts = valid_df.groupby('Date', sort=False).size().rename('Total Menit Log')
        valid_counts = valid_df.groupby('Date', sort=False)[valid_sensors].count()
        
        pct_df = (valid_counts / TOTAL_HARUSNYA_PER_HARI * 100).clip(upper=100.0).round(2)
        pct_df.rename(columns={s: f"{sensor_labels[s]} (%)" for s in valid_sensors}, inplace=True)
        
        total_possible = len(sensor_labels) * TOTAL_HARUSNYA_PER_HARI
        overall_pct = (valid_counts.sum(axis=1) / total_possible * 100).clip(upper=100.0).round(2)
        
        summary_df = pd.concat([counts, pct_df], axis=1)
        summary_df['PERSENTASE TOTAL KESELURUHAN (%)'] = overall_pct
        summary_df.reset_index(inplace=True)

        summary_df['date_temp'] = pd.to_datetime(summary_df['Date'], errors='coerce')
        summary_df.sort_values('date_temp', inplace=True)
        summary_df.drop(columns=['date_temp'], inplace=True, errors='ignore')
        
        if not summary_df.empty:
            avg_row = {
                'Date': 'RATA-RATA & TOTAL BULANAN', 
                'Total Menit Log': int(summary_df['Total Menit Log'].sum())
            }
            num_cols = summary_df.columns.difference(['Date', 'Total Menit Log'])
            avg_row.update(summary_df[num_cols].mean().round(2).to_dict())
            summary_df = pd.concat([summary_df, pd.DataFrame([avg_row])], ignore_index=True)
            
        return summary_df

    @staticmethod
    def simpan_ke_excel_bytes(df_data, df_summary, data_sheet_name, sensor_labels):
        output = io.BytesIO()
        
        dim_date = EnginePerapihData.generate_dim_date(df_data)
        dim_time = EnginePerapihData.generate_dim_time()
        dim_sensor = EnginePerapihData.generate_dim_sensor(sensor_labels)

        fact_df = df_data.drop(columns=['Date', 'Time'], errors='ignore')
        key_cols = ['PK_Datetime', 'DateKey', 'TimeKey']
        other_cols = [c for c in fact_df.columns if c not in key_cols]
        fact_df = fact_df[key_cols + other_cols]

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            fact_df.to_excel(writer, index=False, sheet_name=data_sheet_name)
            df_summary.to_excel(writer, index=False, sheet_name="Rangkuman_Ketersediaan")
            dim_date.to_excel(writer, index=False, sheet_name="Dim_Date")
            dim_time.to_excel(writer, index=False, sheet_name="Dim_Time")
            dim_sensor.to_excel(writer, index=False, sheet_name="Dim_Sensor")
            
            font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
            font_data = Font(name='Arial', size=10)
            fill_zebra = PatternFill(start_color='F2F5F9', end_color='F2F5F9', fill_type='solid')
            align_center = Alignment(horizontal='center', vertical='center')
            side_thin = Side(style='thin', color='D9D9D9')
            border_thin = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)

            for sheetname in [data_sheet_name, "Rangkuman_Ketersediaan", "Dim_Date", "Dim_Time", "Dim_Sensor"]:
                ws = writer.sheets[sheetname]
                ws.freeze_panes = 'A2'
                ws.row_dimensions[1].height = 28
                
                for cell in ws[1]:
                    cell.font, cell.fill, cell.alignment, cell.border = font_header, fill_header, align_header, border_thin

                for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
                    ws.row_dimensions[r_idx].height = 20
                    is_even = (r_idx % 2 == 0)
                    for cell in row:
                        cell.font = font_data
                        cell.border = border_thin
                        if is_even: cell.fill = fill_zebra
                        cell.alignment = align_center

                for col in ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 50)

        return output.getvalue()

    @classmethod
    @st.cache_data(show_spinner=False)
    def baca_database_fdb(cls, fdb_bytes, kolom_aws_resmi):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".fdb") as tmp:
            tmp.write(fdb_bytes)
            tmp_path = tmp.name

        conn = None
        try:
            try:
                import fdb
                conn = fdb.connect(dsn=tmp_path, user='sysdba', password='masterkey')
            except Exception:
                from firebird.driver import connect
                conn = connect(tmp_path, user='sysdba', password='masterkey')
                
            cursor = conn.cursor()
            cursor.execute("SELECT RDB$RELATION_NAME FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG = 0 AND RDB$VIEW_BLR IS NULL")
            all_tables = [row[0].strip().upper() for row in cursor.fetchall()]
            
            priority_order = ['T1MN', 'TBRUT', 'T1H']
            sorted_tables = [t for t in priority_order if t in all_tables] + [t for t in all_tables if t not in priority_order]
            
            extracted_dfs = []
            for target_table in sorted_tables:
                try:
                    cursor.execute(f'SELECT * FROM "{target_table}"')
                    df_tbl = pd.DataFrame(cursor.fetchall(), columns=[f[0] for f in cursor.description])
                    if not df_tbl.empty:
                        df_mapped = pd.DataFrame(index=df_tbl.index)
                        dt_col = next((c for c in df_tbl.columns if c.upper() in ['DATACQ', 'DATETIME', 'DATE_TIME', 'DATE']), None)
                        if dt_col:
                            dt_parsed = pd.to_datetime(df_tbl[dt_col], errors='coerce')
                            df_mapped['Date'] = dt_parsed.dt.strftime('%Y-%m-%d')
                            df_mapped['Time'] = dt_parsed.dt.strftime('%H:%M:00')
                            for c in kolom_aws_resmi:
                                if c not in ['Date', 'Time']:
                                    df_mapped[c] = df_tbl[c] if c in df_tbl.columns else ''
                            extracted_dfs.append(df_mapped[kolom_aws_resmi])
                            if target_table in ['T1MN', 'TBRUT']: 
                                break
                except Exception:
                    continue
            
            return pd.concat(extracted_dfs, ignore_index=True) if extracted_dfs else pd.DataFrame(columns=kolom_aws_resmi)
        finally:
            if conn: 
                conn.close()
            if os.path.exists(tmp_path): 
                os.remove(tmp_path)


# --- INTERFASE PENGGUNA (STREAMLIT UI) ---
st.title("🌤️ Pengolah Data ALOPTAMA")
st.caption("Aplikasi Perapih Data AWOS & AWS Strengthening - By Luqmanul Hakim, S.Tr")

tab_awos, tab_fdb, tab_aws = st.tabs([
    "1. 📊 Panel AWOS Universal", 
    "2. ⚡ Ekstrak FDB ke CSV", 
    "3. 🛰️ Panel AWS Strengthening"
])

# --- TAB 1: AWOS UNIVERSAL ---
with tab_awos:
    st.subheader("Pengolahan Data AWOS (.CSV / .XLSX)")
    files_awos = st.file_uploader("Upload File Mentah AWOS", type=['csv', 'xlsx', 'xls'], accept_multiple_files=True, key="awos_new")
    
    if st.button("Proses Data AWOS 🚀", key="btn_awos"):
        if not files_awos:
            st.error("Silakan upload AWOS!")
        else:
            with st.spinner("Sedang diproses..."):
                df_all = EnginePerapihData.load_files_to_dataframe(files_awos)
                df_clean = EnginePerapihData.normalize_dataframe(df_all)
                df_summary = EnginePerapihData.buat_rangkuman_per_sensor(df_clean, EnginePerapihData.UNIVERSAL_LABELS)
                excel_bytes = EnginePerapihData.simpan_ke_excel_bytes(
                    df_clean, df_summary, "Fact_AWOS_Data", EnginePerapihData.UNIVERSAL_LABELS
                )
                
                st.success("✅ Data AWOS Berhasil Diproses!")
                st.download_button(
                    label="⬇️ Download Excel AWOS",
                    data=excel_bytes,
                    file_name="AWOS_Clean.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# --- TAB 2: FDB TO CSV ---
with tab_fdb:
    st.subheader("Ekstrak Database Firebird AWS (.FDB) Kilat ke CSV (UTC)")
    fdb_file = st.file_uploader("Upload File Database Firebird (.FDB)", type=['fdb'], key="fdb_single")
    
    if st.button("⚡ Ekstrak Kilat ke CSV", key="btn_fdb"):
        if not fdb_file:
            st.error("Upload file .FDB terlebih dahulu!")
        else:
            with st.spinner("Mengekstrak tabel database FDB..."):
                df_fdb = EnginePerapihData.baca_database_fdb(fdb_file.getbuffer(), EnginePerapihData.KOLOM_AWS_RESMI)
                df_clean = EnginePerapihData.normalize_dataframe(df_fdb)
                
                csv_bytes = df_clean.to_csv(index=False).encode('utf-8')
                st.success("⚡ Ekstraksi FDB Selesai!")
                st.download_button(
                    label="⬇️ Download Hasil CSV",
                    data=csv_bytes,
                    file_name="AWS_Ekstrak_FDB_UTC.csv",
                    mime="text/csv"
                )

# --- TAB 3: AWS ---
with tab_aws:
    st.subheader("Pengolahan Data AWS (.XLSX / .CSV / .TXT / .FDB)")
    files_aws = st.file_uploader("Upload File AWS (Text/Excel/CSV/FDB)", type=['xlsx', 'xls', 'csv', 'txt', 'fdb'], accept_multiple_files=True, key="aws_new")
    
    if st.button("Proses Data AWS 🚀", key="btn_aws"):
        if not files_aws:
            st.error("Silakan upload file AWS!")
        else:
            with st.spinner("Sedang diproses..."):
                df_all = EnginePerapihData.load_files_to_dataframe(files_aws)
                df_clean = EnginePerapihData.normalize_dataframe(df_all)
                df_summary = EnginePerapihData.buat_rangkuman_per_sensor(df_clean, EnginePerapihData.UNIVERSAL_LABELS)
                excel_bytes = EnginePerapihData.simpan_ke_excel_bytes(
                    df_clean, df_summary, "Fact_AWS_Data", EnginePerapihData.UNIVERSAL_LABELS
                )
                
                st.success("✅ Data AWS Berhasil Diproses!")
                st.download_button(
                    label="⬇️ Download Excel AWS",
                    data=excel_bytes,
                    file_name="AWS_Clean.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
