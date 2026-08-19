import streamlit as st
import pandas as pd
import re
import io
import os
import calendar
import zipfile
import openpyxl
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.worksheet.pagebreak import Break

# ==========================================
# ===== KONFIGURASI HALAMAN UTAMA ==========
# ==========================================
st.set_page_config(page_title="BMKG Data Generator", layout="centered", page_icon="🌤️")

# --- DICTIONARY BULAN BAHASA INDONESIA (KAPITAL) ---
BULAN_INDO = {
    1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL",
    5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
    9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER"
}

def safe_set_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell): 
        cell.value = value

# ==========================================
# ===== METAR & SPECI PARSER FUNCTIONS =====
# ==========================================
def parse_metar_speci(sandi_str):
    if pd.isna(sandi_str):
        return None
    sandi_str = sandi_str.replace('\n', ' ').replace('\r', '').strip()
    
    report_type = None
    if 'METAR' in sandi_str:
        report_type = 'METAR'
    elif 'SPECI' in sandi_str:
        report_type = 'SPECI'
    else:
        return None
        
    start_idx = sandi_str.find(report_type)
    core_str = sandi_str[start_idx:].replace('=', '').strip()
    tokens = core_str.split()
    
    header_title, loc, time_str, wind, vis, wx, cloud, t_dp, qnh, rmk = report_type, "NIL", "NIL", "NIL", "NIL", "NIL", "NIL", "NIL", "NIL", "NOSIG"
    
    is_cor = False
    cc_type = ""
    remaining_tokens = []
    
    for t in tokens:
        if t in ['METAR', 'SPECI']: continue
        elif t == 'COR':
            is_cor = True
            continue
        elif re.match(r'^CC[A-Z]$', t):
            cc_type = t.upper()
            continue
        elif re.match(r'^[A-Z]{4}$', t) and loc == "NIL":
            loc = t
            continue
        elif re.match(r'^\d{6}Z$', t) and time_str == "NIL":
            time_str = t
            continue
        else:
            remaining_tokens.append(t)
            
    if 'CAVOK' in remaining_tokens:
        vis = 'CAVOK'
        wx = ''      
        cloud = ''   
        for t in remaining_tokens:
            if re.match(r'^\d{5}(G\d{2})?KT$', t) or re.match(r'^VRB\d{2}KT$', t) or t == '00000KT': wind = t
            elif re.match(r'^\d{2}/\d{2}$', t) or re.match(r'^M\d{2}/\d{2}$', t): t_dp = t.replace('/', ' / ')
            elif re.match(r'^Q\d{4}$', t): qnh = t
            elif t in ['NOSIG', 'TEMPO', 'BECMG']: rmk = t
    else:
        cloud_list = []
        for t in remaining_tokens:
            if re.match(r'^\d{5}(G\d{2})?KT$', t) or re.match(r'^VRB\d{2}KT$', t) or t == '00000KT': wind = t
            elif re.match(r'^\d{4}$', t): vis = t
            elif t in ['RA', 'DZ', 'SHRA', 'TSRA', 'TS', 'BR', 'HZ', 'FG', '-RA', '+RA', 'VCTS', '+TSRA', '-TSRA']: wx = t
            elif re.match(r'^(FEW|SCT|BKN|OVC)\d{3}(CB|TCU)?$', t) or t in ['NSC', 'SKC', 'CLR']: cloud_list.append(t)
            elif re.match(r'^\d{2}/\d{2}$', t) or re.match(r'^M\d{2}/\d{2}$', t): t_dp = t.replace('/', ' / ')
            elif re.match(r'^Q\d{4}$', t): qnh = t
            elif t in ['NOSIG', 'TEMPO', 'BECMG']: rmk = t
        if cloud_list:
            cloud = " ".join(cloud_list)
            
    return [header_title, loc, time_str, wind, vis, wx, cloud, t_dp, qnh, rmk, is_cor, cc_type]

def calculate_priority(row):
    score = 0
    if row['is_cor']: score = 1
    if row['cc_type'] and len(row['cc_type']) == 3:
        char_code = ord(row['cc_type'][2]) - ord('A')
        score = max(score, 2 + char_code)
    return score

def check_is_thunderstorm(wx, cloud):
    wx_str = str(wx).upper() if pd.notna(wx) else ""
    cloud_str = str(cloud).upper() if pd.notna(cloud) else ""
    has_ts = any(code in wx_str for code in ['TS', 'TSRA', 'VCTS'])
    has_cb = 'CB' in cloud_str
    return has_ts or has_cb

def get_cloud_category(cloud_str, vis_str):
    if vis_str == 'CAVOK' or pd.isna(cloud_str) or str(cloud_str).strip() in ['', 'NIL', 'SKC', 'CLR', 'NSC']:
        return 'CERAH'
    
    c_str = str(cloud_str).upper()
    if 'OVC' in c_str:
        return 'BERAWAN_BANYAK'
    elif 'BKN' in c_str:
        return 'BERAWAN'
    elif 'SCT' in c_str or 'FEW' in c_str:
        return 'BERAWAN_SEBAGIAN'
    else:
        return 'CERAH'

def generate_excel_bytes_metar_speci_fallback(df_clean, report_type="METAR"):
    buffer = io.BytesIO()
    headers = ['TYPE', 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK', 'datetime']
    df_export = df_clean[headers].copy()
    df_export.rename(columns={'TYPE': report_type}, inplace=True)
    df_export['datetime'] = df_export['datetime'].dt.strftime('%Y-%m-%d %H:%M:%S')
    
    sheet_title = f"Rekap {report_type}"
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name=sheet_title, index=False)
        worksheet = writer.sheets[sheet_title]
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        align_center, align_left = Alignment(horizontal='center', vertical='center'), Alignment(horizontal='left', vertical='center')
        thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
        
        for cell in worksheet[1]:
            cell.font, cell.fill, cell.alignment = header_font, header_fill, align_center
            
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border, cell.font = thin_border, Font(name='Segoe UI', size=10)
                col_header = worksheet.cell(row=1, column=cell.column).value
                cell.alignment = align_center if col_header in [report_type, 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'T/DP', 'QNH'] else align_left
        
        for col in worksheet.columns:
            max_len = max([len(str(cell.value)) for cell in col if cell.value] + [0])
            worksheet.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 11)
            
    buffer.seek(0)
    return buffer

# =================================================================================================
# ===== EXCEL TEMPLATE INSERTER (METAR: HARIAN 1 HALAMAN LENGKAP KOP) =============================
# =================================================================================================
def generate_excel_from_template_daily(df_clean, report_type="METAR", template_path="TEMPLATE METAR_3.xlsx", logo_path="logo_bmkg.png"):
    if not os.path.exists(template_path):
        for alt in ["TEMPLATE METAR.xlsx", "TEMPLATE METAR_2.xlsx", "TEMPLATE METAR_3.xlsx"]:
            if os.path.exists(alt):
                template_path = alt
                break

    buffer = io.BytesIO()
    wb = openpyxl.load_workbook(template_path)
    
    if "Template" in wb.sheetnames:
        ws_template = wb["Template"]
    elif "TEMPLATE" in wb.sheetnames:
        ws_template = wb["TEMPLATE"]
    else:
        ws_template = wb.active

    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    font_kop_bold = Font(name='Arial', size=11, bold=True)
    font_tbl_header = Font(name='Arial', size=9, bold=True)
    font_data = Font(name='Arial', size=10) 
    
    tbl_header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    headers_table = [report_type, 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK']

    df_clean['year'] = df_clean['datetime'].dt.year
    df_clean['month'] = df_clean['datetime'].dt.month
    df_clean['day'] = df_clean['datetime'].dt.day
    df_clean['hour'] = df_clean['datetime'].dt.hour
    
    grouped_months = df_clean.groupby(['year', 'month'])
    
    for (year, month), month_group in grouped_months:
        nama_bulan = BULAN_INDO[month]
        sheet_name = f"{nama_bulan}"
        
        ws = wb.create_sheet(title=sheet_name)
        
        col_widths = [10.5, 8.5, 10.5, 9.8, 9.0, 7.1, 15.5, 10.0, 9.2, 12.0]
        for c_i, w_val in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_i)].width = w_val

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        
        ws.page_margins.left = 1.0   
        ws.page_margins.right = 0.2  
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.print_options.horizontalCentered = True 
        
        num_days = calendar.monthrange(year, month)[1]
        
        if hasattr(ws, 'row_breaks') and hasattr(ws.row_breaks, 'brk'):
            ws.row_breaks.brk = []

        current_row = 1

        for d in range(1, num_days + 1):
            start_row = current_row
            
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
            c1 = ws.cell(row=start_row, column=1, value="BALAI BESAR METEOROLOGI KLIMATOLOGI DAN GEOFISIKA WILAYAH III")
            c1.font = font_kop_bold; c1.alignment = align_center; ws.row_dimensions[start_row].height = 18

            ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=10)
            c2 = ws.cell(row=start_row+1, column=1, value="STASIUN METEOROLOGI UMBU MEHANG KUNDA")
            c2.font = font_kop_bold; c2.alignment = align_center; ws.row_dimensions[start_row+1].height = 18

            ws.merge_cells(start_row=start_row+2, start_column=1, end_row=start_row+2, end_column=10)
            c3 = ws.cell(row=start_row+2, column=1, value="JL. ADI SUCIPTO NO.3")
            c3.font = font_kop_bold; c3.alignment = align_center; ws.row_dimensions[start_row+2].height = 18

            if os.path.exists(logo_path):
                try:
                    img = OpenpyxlImage(logo_path)
                    img.width = 45; img.height = 45
                    ws.add_image(img, f'A{start_row}')
                except:
                    pass

            ws.merge_cells(start_row=start_row+3, start_column=3, end_row=start_row+3, end_column=6)
            c_label = ws.cell(row=start_row+3, column=3, value=f"REKAP DATA {report_type} :")
            c_label.font = font_kop_bold; c_label.alignment = align_right; ws.row_dimensions[start_row+3].height = 18

            ws.merge_cells(start_row=start_row+3, start_column=7, end_row=start_row+3, end_column=9)
            c_date = ws.cell(row=start_row+3, column=7, value=f"{d:02d} {nama_bulan} {year}")
            c_date.font = font_kop_bold; c_date.alignment = align_left

            ws.row_dimensions[start_row+4].height = 8

            for c_i, h_text in enumerate(headers_table, start=1):
                ws.merge_cells(start_row=start_row+5, start_column=c_i, end_row=start_row+6, end_column=c_i)
                cell_h = ws.cell(row=start_row+5, column=c_i, value=h_text)
                cell_h.font = font_tbl_header; cell_h.fill = tbl_header_fill; cell_h.alignment = align_center
                for r_sub in range(start_row+5, start_row+7):
                    ws.cell(row=r_sub, column=c_i).border = thin_border
            ws.row_dimensions[start_row+5].height = 15
            ws.row_dimensions[start_row+6].height = 15

            day_data = month_group[month_group['day'] == d]
            data_row_idx = start_row + 7
            
            day_map = {r['hour']: r for _, r in day_data.iterrows()}
            for h in range(24):
                if h in day_map:
                    row = day_map[h]
                    metar_type = str(row['TYPE']) if pd.notna(row['TYPE']) else 'METAR'
                    loc = str(row['LOC']) if pd.notna(row['LOC']) else 'WATU'
                    time_str = str(row['TIME']) if pd.notna(row['TIME']) else f"{d:02d}{h:02d}00Z"
                    wind = str(row['WIND']) if pd.notna(row['WIND']) else ''
                    vis = str(row['VIS']) if pd.notna(row['VIS']) else ''
                    wx = str(row['WX']) if pd.notna(row['WX']) else ''
                    cloud = str(row['CLOUD']) if pd.notna(row['CLOUD']) else ''
                    t_dp = str(row['T/DP']) if pd.notna(row['T/DP']) else ''
                    qnh = str(row['QNH']) if pd.notna(row['QNH']) else ''
                    rmk = str(row['RMK']) if pd.notna(row['RMK']) else ''
                    if vis == 'CAVOK':
                        wx = ''; cloud = ''
                    vals = [metar_type, loc, time_str, wind, int(vis) if vis.isdigit() else vis, wx, cloud, t_dp, qnh, rmk]
                else:
                    vals = ['METAR', 'WATU', f"{d:02d}{h:02d}00Z", 'NIL', 'NIL', 'NIL', 'NIL', 'NIL', 'NIL', 'NOSIG']

                for col_idx, val in enumerate(vals, start=1):
                    cell = ws.cell(row=data_row_idx, column=col_idx, value=val)
                    cell.font = font_data; cell.alignment = align_center; cell.border = thin_border
                
                ws.row_dimensions[data_row_idx].height = 23.5 
                data_row_idx += 1
            
            if d < num_days:
                ws.row_breaks.append(Break(id=data_row_idx - 1))
            
            current_row = data_row_idx

        ws.print_title_rows = None 
        
    wb.remove(ws_template)
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# =================================================================================================
# ===== EXCEL TEMPLATE INSERTER (SPECI: BULANAN BERSAMBUNG 1 KOP PENUH) ============================
# =================================================================================================
def generate_excel_from_template_speci(df_clean, report_type="SPECI", template_path="TEMPLATE METAR_3.xlsx", logo_path="logo_bmkg.png"):
    if not os.path.exists(template_path):
        for alt in ["TEMPLATE METAR.xlsx", "TEMPLATE METAR_2.xlsx", "TEMPLATE METAR_3.xlsx"]:
            if os.path.exists(alt):
                template_path = alt
                break

    buffer = io.BytesIO()
    wb = openpyxl.load_workbook(template_path)
    
    if "Template" in wb.sheetnames:
        ws_template = wb["Template"]
    elif "TEMPLATE" in wb.sheetnames:
        ws_template = wb["TEMPLATE"]
    else:
        ws_template = wb.active

    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    font_kop_bold = Font(name='Arial', size=11, bold=True)
    font_tbl_header = Font(name='Arial', size=9, bold=True)
    font_data = Font(name='Arial', size=10)
    
    tbl_header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    headers_table = [report_type, 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK']

    df_clean['year'] = df_clean['datetime'].dt.year
    df_clean['month'] = df_clean['datetime'].dt.month

    grouped_months = df_clean.groupby(['year', 'month'])
    
    for (year, month), month_group in grouped_months:
        nama_bulan = BULAN_INDO[month]
        sheet_name = f"SPECI {nama_bulan[:3]}"
        
        ws = wb.create_sheet(title=sheet_name)
        
        col_widths = [10.5, 8.5, 10.5, 9.8, 9.0, 7.1, 15.5, 10.0, 9.2, 12.0]
        for c_i, w_val in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_i)].width = w_val

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        ws.page_margins.left = 1.0   
        ws.page_margins.right = 0.2  
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.print_options.horizontalCentered = True 
        
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        c1 = ws.cell(row=1, column=1, value="BALAI BESAR METEOROLOGI KLIMATOLOGI DAN GEOFISIKA WILAYAH III")
        c1.font = font_kop_bold; c1.alignment = align_center; ws.row_dimensions[1].height = 18

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
        c2 = ws.cell(row=2, column=1, value="STASIUN METEOROLOGI UMBU MEHANG KUNDA")
        c2.font = font_kop_bold; c2.alignment = align_center; ws.row_dimensions[2].height = 18

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
        c3 = ws.cell(row=3, column=1, value="JL. ADI SUCIPTO NO.3")
        c3.font = font_kop_bold; c3.alignment = align_center; ws.row_dimensions[3].height = 18

        if os.path.exists(logo_path):
            try:
                img = OpenpyxlImage(logo_path)
                img.width = 45; img.height = 45
                ws.add_image(img, 'A1')
            except:
                pass

        ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=6)
        c_label = ws.cell(row=4, column=3, value=f"REKAP DATA {report_type} :")
        c_label.font = font_kop_bold; c_label.alignment = align_right; ws.row_dimensions[4].height = 18

        ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=9)
        c_date = ws.cell(row=4, column=7, value=f"{nama_bulan} {year}")
        c_date.font = font_kop_bold; c_date.alignment = align_left
        ws.row_dimensions[5].height = 8

        for c_i, h_text in enumerate(headers_table, start=1):
            ws.merge_cells(start_row=6, start_column=c_i, end_row=7, end_column=c_i)
            cell_h = ws.cell(row=6, column=c_i, value=h_text)
            cell_h.font = font_tbl_header; cell_h.fill = tbl_header_fill; cell_h.alignment = align_center
            for r_sub in range(6, 8):
                ws.cell(row=r_sub, column=c_i).border = thin_border
        ws.row_dimensions[6].height = 15; ws.row_dimensions[7].height = 15
        
        ws.print_title_rows = '$6:$7'

        data_row_idx = 8
        month_group = month_group.sort_values('datetime')
        
        if month_group.empty:
            vals = ['SPECI', 'WATU', f"NIL", 'NIL', 'NIL', 'NIL', 'NIL', 'NIL', 'NIL', 'NOSIG']
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=data_row_idx, column=col_idx, value=val)
                cell.font = font_data; cell.alignment = align_center; cell.border = thin_border
            ws.row_dimensions[data_row_idx].height = 18
        else:
            for _, row in month_group.iterrows():
                speci_type = str(row['TYPE']) if pd.notna(row['TYPE']) else 'SPECI'
                loc = str(row['LOC']) if pd.notna(row['LOC']) else 'WATU'
                time_str = str(row['TIME']) if pd.notna(row['TIME']) else ''
                wind = str(row['WIND']) if pd.notna(row['WIND']) else ''
                vis = str(row['VIS']) if pd.notna(row['VIS']) else ''
                wx = str(row['WX']) if pd.notna(row['WX']) else ''
                cloud = str(row['CLOUD']) if pd.notna(row['CLOUD']) else ''
                t_dp = str(row['T/DP']) if pd.notna(row['T/DP']) else ''
                qnh = str(row['QNH']) if pd.notna(row['QNH']) else ''
                rmk = str(row['RMK']) if pd.notna(row['RMK']) else ''

                if vis == 'CAVOK':
                    wx = ''; cloud = ''
                    
                vals = [speci_type, loc, time_str, wind, int(vis) if vis.isdigit() else vis, wx, cloud, t_dp, qnh, rmk]
                for col_idx, val in enumerate(vals, start=1):
                    cell = ws.cell(row=data_row_idx, column=col_idx, value=val)
                    cell.font = font_data; cell.alignment = align_center; cell.border = thin_border
                ws.row_dimensions[data_row_idx].height = 18
                data_row_idx += 1

    wb.remove(ws_template)
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# =================================================================================================
# ===== EXCEL TEMPLATE INSERTER (WXREV: BERDASARKAN TEMPLATE WXREV.XLSX) ==========================
# =================================================================================================
def generate_excel_from_template_wxrev(df_wxrev, template_path="TEMPLATE WXREV.xlsx"):
    if not os.path.exists(template_path):
        template_path = "TEMPLATE WXREV.xlsx"

    buffer = io.BytesIO()
    wb = openpyxl.load_workbook(template_path)
    
    if "Sheet1" in wb.sheetnames:
        ws_template = wb["Sheet1"]
    else:
        ws_template = wb.active

    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    align_center = Alignment(horizontal='center', vertical='center')
    font_data = Font(name='Arial', size=9)

    df_wxrev['year'] = df_wxrev['datetime'].dt.year
    df_wxrev['month'] = df_wxrev['datetime'].dt.month
    df_wxrev['day'] = df_wxrev['datetime'].dt.day

    grouped_months = df_wxrev.groupby(['year', 'month'])
    
    for (year, month), month_group in grouped_months:
        nama_bulan = BULAN_INDO[month]
        sheet_name = f"WXREV {nama_bulan[:3]}"
        
        ws = wb.copy_worksheet(ws_template)
        ws.title = sheet_name
        
        # 1. BULAN DI CELL F4
        ws['F4'].value = f"{nama_bulan} {year}"
        
        # Tanggal TTD di H40
        num_days = calendar.monthrange(year, month)[1]
        ws['H40'].value = f"WAINGAPU, {num_days:02d} {nama_bulan} {year}"
        
        day_map = {r['day']: r for _, r in month_group.iterrows()}
        
        # 2. ISU SEL DARI BARIS 8 SAMPAI BARIS 38 (C8:J38)
        for d in range(1, 32):
            target_row = 7 + d # Hari 1 -> Row 8, Hari 31 -> Row 38
            
            if d <= num_days and d in day_map:
                row = day_map[d]
                mmyyg = str(row['MMYYGp']) if pd.notna(row['MMYYGp']) else ''
                iiiii = str(row['IIiii']) if pd.notna(row['IIiii']) else ''
                att = str(row['atTxTxTnTn']) if pd.notna(row['atTxTxTnTn']) else ''
                app = str(row['apPxPxPnPn']) if pd.notna(row['apPxPxPnPn']) else ''
                auu = str(row['auUxUxUnUn']) if pd.notna(row['auUxUxUnUn']) else ''
                arr = str(row['arRRRR']) if pd.notna(row['arRRRR']) else ''
                rdrd = str(row['rDrDdfmfm_1']) if pd.notna(row['rDrDdfmfm_1']) else ''
                
                # Jam kirim (Format HH:MM WITA)
                dt_send = row['datetime']
                jam_kirim = dt_send.strftime('%H:%M WITA') if pd.notna(dt_send) else ''
                
                vals = [mmyyg, iiiii, att, app, auu, arr, rdrd, jam_kirim]
            else:
                vals = ['-', '-', '-', '-', '-', '-', '-', '-'] if d <= num_days else ['', '', '', '', '', '', '', '']
                
            col_indices = [3, 4, 5, 6, 7, 8, 9, 10] # Kolom C, D, E, F, G, H, I, J
            for col_idx, val in zip(col_indices, vals):
                cell = ws.cell(row=target_row, column=col_idx, value=val)
                cell.font = font_data
                cell.alignment = align_center
                cell.border = thin_border

    wb.remove(ws_template)
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def parse_wxrev(sandi_str):
    if pd.isna(sandi_str):
        return None
    sandi_str = sandi_str.replace('\n', ' ').replace('\r', '').replace('=', '').strip()
    
    start_idx = sandi_str.find('WXREV')
    if start_idx == -1:
        return None
        
    wxrev_core = sandi_str[start_idx:].strip()
    tokens = wxrev_core.split()
    
    if len(tokens) < 7: 
        return None
        
    mmyygp = tokens[1] if len(tokens) > 1 else ""
    tgl = mmyygp[2:4] if len(mmyygp) >= 4 else ""
    iiiii = tokens[2] if len(tokens) > 2 else ""
    att = tokens[3] if len(tokens) > 3 else ""
    app = tokens[4] if len(tokens) > 4 else ""
    auu = tokens[5] if len(tokens) > 5 else ""
    arr = tokens[6] if len(tokens) > 6 else ""
    rdrd1 = tokens[7] if len(tokens) > 7 else ""
    rdrd2 = tokens[8] if len(tokens) > 8 else ""
    
    return [tgl, mmyygp, iiiii, att, app, auu, arr, rdrd1, rdrd2]

def generate_pdf_bytes_thunderstorm(df_clean, station_name, kepala_nama, kepala_nip):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TSTitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, leading=16, alignment=1)
    meta_style = ParagraphStyle('TSMeta', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=14, alignment=1)
    text_style = ParagraphStyle('TSText', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=11)
    
    tahun = df_clean['datetime'].dt.year.iloc[0]
    
    story.append(Paragraph("DATA THUNDERSTORM", title_style))
    story.append(Paragraph(f"{station_name.upper()}", meta_style))
    story.append(Paragraph(f"TAHUN: {tahun}", meta_style))
    story.append(Spacer(1, 10))
    
    matrix = {m: {d: "" for d in range(1, 32)} for m in range(1, 13)}
    
    df_clean['day'] = df_clean['datetime'].dt.day
    df_clean['month'] = df_clean['datetime'].dt.month
    df_clean['is_ts'] = df_clean.apply(lambda r: check_is_thunderstorm(r['WX'], r['CLOUD']), axis=1)
    
    observed_dates = df_clean[['month', 'day']].drop_duplicates()
    for _, r in observed_dates.iterrows():
        matrix[r['month']][r['day']] = "O"
        
    ts_dates = df_clean[df_clean['is_ts'] == True][['month', 'day']].drop_duplicates()
    for _, r in ts_dates.iterrows():
        matrix[r['month']][r['day']] = "X"
        
    header_row = ['TANGGAL\nBULAN'] + [str(d) for d in range(1, 32)]
    table_data = [header_row]
    
    for m in range(1, 13):
        row = [BULAN_INDO[m]]
        for d in range(1, 32):
            row.append(matrix[m][d])
        table_data.append(row)
        
    col_widths = [85] + [22] * 31
    ts_table = Table(table_data, colWidths=col_widths)
    
    ts_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3), ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'), ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
    ]))
    
    story.append(ts_table)
    story.append(Spacer(1, 15))
    
    tgl_sekarang = datetime.now()
    tgl_ttd = f"WAINGAPU, {tgl_sekarang.day} {BULAN_INDO[tgl_sekarang.month]} {tgl_sekarang.year}"
    
    ket_text = """<b>KETERANGAN:</b><br/>
    <b>X</b> : Ada satu atau lebih kilat / thunderstorm dalam sandi wwW1W2.<br/>
    <b>O</b> : Tidak ada kilat / thunderstorm dalam sandi wwW1W2.<br/>
    &nbsp;&nbsp;&nbsp;&nbsp;: Tidak ada pengamatan.
    """
    
    ttd_text = f"""{tgl_ttd}<br/>
    <b>KEPALA STASIUN</b><br/>
    <b>BADAN METEOROLOGI KLIMATOLOGI DAN GEOFISIKA</b><br/>
    <b>{station_name.upper()}</b><br/><br/><br/><br/>
    <b><u>{kepala_nama}</u></b>
    """
    
    ket_p = Paragraph(ket_text, text_style)
    ttd_p = Paragraph(ttd_text, ParagraphStyle('TTD', parent=text_style, alignment=1))
    
    footer_table = Table([[ket_p, ttd_p]], colWidths=[400, 367])
    footer_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    
    story.append(KeepTogether(footer_table))
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_bytes_thunderstorm(df_clean, station_name):
    buffer = io.BytesIO()
    tahun = df_clean['datetime'].dt.year.iloc[0]
    
    matrix = {m: {d: "" for d in range(1, 32)} for m in range(1, 13)}
    df_clean['day'] = df_clean['datetime'].dt.day
    df_clean['month'] = df_clean['datetime'].dt.month
    df_clean['is_ts'] = df_clean.apply(lambda r: check_is_thunderstorm(r['WX'], r['CLOUD']), axis=1)
    
    observed_dates = df_clean[['month', 'day']].drop_duplicates()
    for _, r in observed_dates.iterrows():
        matrix[r['month']][r['day']] = "O"
        
    ts_dates = df_clean[df_clean['is_ts'] == True][['month', 'day']].drop_duplicates()
    for _, r in ts_dates.iterrows():
        matrix[r['month']][r['day']] = "X"
        
    rows_data = []
    for m in range(1, 13):
        r_dict = {'BULAN': BULAN_INDO[m]}
        for d in range(1, 32):
            r_dict[str(d)] = matrix[m][d]
        rows_data.append(r_dict)
        
    df_ts = pd.DataFrame(rows_data)
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_ts.to_excel(writer, sheet_name=f'Thunderstorm {tahun}', index=False)
        worksheet = writer.sheets[f'Thunderstorm {tahun}']
        
        header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
        
        for cell in worksheet[1]:
            cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, align_center, thin_border
            
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border, cell.font, cell.alignment = thin_border, Font(name='Segoe UI', size=10), align_center
                
    buffer.seek(0)
    return buffer

def generate_pdf_bytes_perawanan(df_clean, station_name, kepala_nama, kepala_nip, form_code="Klim / PWN/WPU-2001"):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=25, bottomMargin=25)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('PWNTitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=13, alignment=1)
    text_style = ParagraphStyle('PWNText', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=11)
    
    df_clean['year'] = df_clean['datetime'].dt.year
    df_clean['month'] = df_clean['datetime'].dt.month
    df_clean['day'] = df_clean['datetime'].dt.day
    df_clean['cat'] = df_clean.apply(lambda r: get_cloud_category(r['CLOUD'], r['VIS']), axis=1)
    
    grouped_month = df_clean.groupby(['year', 'month'])
    
    for count_month, ((year, month), month_group) in enumerate(grouped_month):
        if count_month > 0:
            story.append(PageBreak())
            
        bulan_str = BULAN_INDO[month]
        num_days = calendar.monthrange(year, month)[1]
        
        story.append(Paragraph("DAFTAR IKHTISAR FREKUENSI PERAWANAN", title_style))
        story.append(Paragraph(f"{station_name.upper()}", title_style))
        story.append(Paragraph(f"BULAN {bulan_str} {year}", title_style))
        story.append(Spacer(1, 10))
        
        header_row = ["TANGGAL", "CERAH\n(N=0)", "BERAWAN SEBAGIAN\n(N=1-3)", "BERAWAN\n(N=4-6)", "BERAWAN BANYAK\n(N=7-8)", "JUMLAH"]
        table_data = [header_row]
        
        tot_cerah = tot_sebagian = tot_berawan = tot_banyak = tot_semua = 0
        
        for d in range(1, num_days + 1):
            df_day = month_group[month_group['day'] == d]
            counts = df_day['cat'].value_counts()
            
            c_cerah = counts.get('CERAH', 0)
            c_sebagian = counts.get('BERAWAN_SEBAGIAN', 0)
            c_berawan = counts.get('BERAWAN', 0)
            c_banyak = counts.get('BERAWAN_BANYAK', 0)
            c_total = len(df_day)
            
            tot_cerah += c_cerah
            tot_sebagian += c_sebagian
            tot_berawan += c_berawan
            tot_banyak += c_banyak
            tot_semua += c_total
            
            table_data.append([
                f"{d:02d}", str(c_cerah) if c_cerah > 0 else "-", str(c_sebagian) if c_sebagian > 0 else "-",
                str(c_berawan) if c_berawan > 0 else "-", str(c_banyak) if c_banyak > 0 else "-", str(c_total) if c_total > 0 else "-"
            ])
            
        table_data.append(["JUMLAH", str(tot_cerah), str(tot_sebagian), str(tot_berawan), str(tot_banyak), str(tot_semua)])
        
        pct_c = f"{round((tot_cerah/tot_semua)*100)}%" if tot_semua > 0 else "0%"
        pct_s = f"{round((tot_sebagian/tot_semua)*100)}%" if tot_semua > 0 else "0%"
        pct_bw = f"{round((tot_berawan/tot_semua)*100)}%" if tot_semua > 0 else "0%"
        pct_by = f"{round((tot_banyak/tot_semua)*100)}%" if tot_semua > 0 else "0%"
        
        table_data.append(["% - TAGE", pct_c, pct_s, pct_bw, pct_by, "100%" if tot_semua > 0 else "0%"])
        
        col_widths = [60, 90, 120, 100, 110, 60]
        pwn_table = Table(table_data, colWidths=col_widths)
        pwn_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2), ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'), ('BACKGROUND', (0, -2), (-1, -1), colors.whitesmoke),
        ]))
        
        story.append(pwn_table)
        story.append(Spacer(1, 15))
        
        tgl_sekarang = datetime.now()
        tgl_ttd = f"WAINGAPU, {tgl_sekarang.day:02d} {BULAN_INDO[tgl_sekarang.month]} {tgl_sekarang.year}"
        
        form_text = f"FORM: {form_code}"
        nip_str = f"<br/>NIP. {kepala_nip}" if kepala_nip else ""
        ttd_text = f"""{tgl_ttd}<br/>
        <b>KEPALA STASIUN METEOROLOGI</b><br/>
        <b>{station_name.upper()}</b><br/><br/><br/><br/>
        <b><u>{kepala_nama}</u></b>{nip_str}
        """
        
        form_p = Paragraph(form_text, text_style)
        ttd_p = Paragraph(ttd_text, ParagraphStyle('TTDPWN', parent=text_style, alignment=1))
        
        footer_table = Table([[form_p, ttd_p]], colWidths=[200, 340])
        footer_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        story.append(KeepTogether(footer_table))

    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_bytes_perawanan(df_clean, station_name):
    buffer = io.BytesIO()
    df_clean['year'] = df_clean['datetime'].dt.year
    df_clean['month'] = df_clean['datetime'].dt.month
    df_clean['day'] = df_clean['datetime'].dt.day
    df_clean['cat'] = df_clean.apply(lambda r: get_cloud_category(r['CLOUD'], r['VIS']), axis=1)
    
    grouped_month = df_clean.groupby(['year', 'month'])
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        for (year, month), month_group in grouped_month:
            bulan_str = BULAN_INDO[month]
            num_days = calendar.monthrange(year, month)[1]
            rows_data = []
            tot_cerah = tot_sebagian = tot_berawan = tot_banyak = tot_semua = 0
            
            for d in range(1, num_days + 1):
                df_day = month_group[month_group['day'] == d]
                counts = df_day['cat'].value_counts()
                
                c_cerah = counts.get('CERAH', 0)
                c_sebagian = counts.get('BERAWAN_SEBAGIAN', 0)
                c_berawan = counts.get('BERAWAN', 0)
                c_banyak = counts.get('BERAWAN_BANYAK', 0)
                c_total = len(df_day)
                
                tot_cerah += c_cerah; tot_sebagian += c_sebagian
                tot_berawan += c_berawan; tot_banyak += c_banyak; tot_semua += c_total
                
                rows_data.append({
                    'TANGGAL': f"{d:02d}",
                    'CERAH (N=0)': c_cerah if c_cerah > 0 else "-",
                    'BERAWAN SEBAGIAN (N=1-3)': c_sebagian if c_sebagian > 0 else "-",
                    'BERAWAN (N=4-6)': c_berawan if c_berawan > 0 else "-",
                    'BERAWAN BANYAK (N=7-8)': c_banyak if c_banyak > 0 else "-",
                    'JUMLAH': c_total if c_total > 0 else "-"
                })
                
            rows_data.append({
                'TANGGAL': "JUMLAH", 'CERAH (N=0)': tot_cerah, 'BERAWAN SEBAGIAN (N=1-3)': tot_sebagian,
                'BERAWAN (N=4-6)': tot_berawan, 'BERAWAN BANYAK (N=7-8)': tot_banyak, 'JUMLAH': tot_semua
            })
            
            pct_c = f"{round((tot_cerah/tot_semua)*100)}%" if tot_semua > 0 else "0%"
            pct_s = f"{round((tot_sebagian/tot_semua)*100)}%" if tot_semua > 0 else "0%"
            pct_bw = f"{round((tot_berawan/tot_semua)*100)}%" if tot_semua > 0 else "0%"
            pct_by = f"{round((tot_banyak/tot_semua)*100)}%" if tot_semua > 0 else "0%"
            
            rows_data.append({
                'TANGGAL': "% - TAGE", 'CERAH (N=0)': pct_c, 'BERAWAN SEBAGIAN (N=1-3)': pct_s,
                'BERAWAN (N=4-6)': pct_bw, 'BERAWAN BANYAK (N=7-8)': pct_by, 'JUMLAH': "100%" if tot_semua > 0 else "0%"
            })
            
            df_pwn = pd.DataFrame(rows_data)
            sheet_name = f"PWN {bulan_str[:3]} {year}"[:31]
            
            df_pwn.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            
            header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            align_center = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
            
            for cell in worksheet[1]:
                cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, align_center, thin_border
                
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border, cell.font, cell.alignment = thin_border, Font(name='Segoe UI', size=10), align_center
                    
            for col in worksheet.columns:
                max_len = max([len(str(cell.value)) for cell in col if cell.value] + [0])
                worksheet.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 12)
            
    buffer.seek(0)
    return buffer

# ==========================================
# ==== PIBAL / RADIOSONDE FUNCTIONS ========
# ==========================================
PENGAMAT_MAP = {
    "Mitra Agritami": "Mitra", "M. Fauzan Fachrurozi": "Fauzan", "Zulqha Ariandi Al Zikri": "Zulqha",    
    "Hakim": "Hakim", "Andreas Yoga": "Andreas", "Imam Zacky Anwar": "Zacky",               
    "Mohammad Hasyim Hanif": "Hanif", "Ni Luh Ayu Agnes": "Ayu", "Yenny Thenu": "Yenny", "Adi": "Adi"
}

def resolve_alasan_and_pengamat(raw_alasan, raw_pengamat, is_norep=False):
    alasan_res = str(raw_alasan).strip() if pd.notna(raw_alasan) and str(raw_alasan).strip().lower() not in ['nan', 'none'] else ""
    raw_p = str(raw_pengamat).strip() if pd.notna(raw_pengamat) and str(raw_pengamat).strip().lower() not in ['nan', 'none'] else ""
    pengamat_res = ""
    if raw_p:
        name_upper = raw_p.upper()
        matched = False
        for full_name, short_name in PENGAMAT_MAP.items():
            if full_name.upper() in name_upper:
                pengamat_res = short_name
                matched = True
                break
        if not matched:
            pengamat_res = raw_p.title()
    if is_norep and not alasan_res:
        alasan_res = "NOREP"
    return alasan_res, pengamat_res

def detect_columns(columns):
    col_alasan = col_pengamat = None
    for c in columns:
        clow = c.lower().strip()
        if any(kw in clow for kw in ['pengamat', 'observer', 'petugas', 'nama', 'user']):
            if not any(kw in clow for kw in ['alasan', 'keterangan', 'penghentian', 'stasiun']):
                col_pengamat = c
                break
    for c in columns:
        clow = c.lower().strip()
        if any(kw in clow for kw in ['alasan', 'reason', 'keterangan', 'ket', 'penghentian', 'henti']):
            if c != col_pengamat:
                col_alasan = c
                break
    return col_alasan, col_pengamat

def extract_valid_value(series):
    valid = series.dropna().astype(str).str.strip()
    valid = valid[~valid.str.lower().isin(['', 'nan', 'none', 'null', '-', 'nat'])]
    if len(valid) > 0:
        return valid.iloc[0]
    return None

def parse_sandi_pibal_complete(sandi_text, surface_dir, surface_speed):
    wind_data = {}
    logs = [] 
    if pd.isna(sandi_text):
        return wind_data, ["Teks sandi kosong (NaN)."]
    text = str(sandi_text).replace('=', ' ')
    if 'NIL' in text:
        return wind_data, []
    wind_data['PERMUKAAN'] = (surface_dir, surface_speed)

    ppaa_data = {}
    if 'PPAA' in text:
        try:
            ppaa_part = text.split('PPAA', 1)[1]
            for end_mark in ['PPBB', 'PPCC', 'PPDD']:
                if end_mark in ppaa_part: ppaa_part = ppaa_part.split(end_mark)[0]
            tokens_a = ppaa_part.split()
            idx = 0
            while idx < len(tokens_a):
                try:
                    token = tokens_a[idx]
                    if token.startswith('55') and len(token) == 5:
                        if idx + 1 < len(tokens_a) and tokens_a[idx+1].startswith('97'):
                            idx += 2; continue
                        indicators = token[2:]
                        level_map = []
                        for ind in indicators:
                            if ind in ['2', '3'] and '5000 feet' not in [l[0] for l in level_map]: level_map.append(('5000 feet', ind))
                            elif ind == '8': level_map.append(('10000 feet', ind))
                            elif ind == '5': level_map.append(('18000 feet', ind))
                            elif ind == '4': level_map.append(('25000 feet', ind))
                            elif ind == '3': level_map.append(('30000 feet', ind))
                            elif ind == '1': level_map.append(('10000 feet' if '5000 feet' in [l[0] for l in level_map] else '18000 feet', ind))
                            elif ind == '6': level_map.append(('45000 feet', ind))
                            elif ind == '7': level_map.append(('50000 feet', ind))
                        for i, (lvl_name, ind) in enumerate(level_map):
                            try:
                                if idx + 1 + i < len(tokens_a):
                                    w_str = tokens_a[idx + 1 + i]
                                    if len(w_str) == 5 and w_str != '/////' and w_str != '77999':
                                        ppaa_data[lvl_name] = (int(w_str[:3]), int(w_str[3:]))
                                    elif w_str != '/////' and w_str != '77999': logs.append(f"Format angin PPAA tidak valid ({w_str}) di level {lvl_name}.")
                            except Exception as e: logs.append(f"Gagal ekstrak angin PPAA level {lvl_name}: {str(e)}")
                        idx += len(level_map)
                except Exception as e: logs.append(f"Gagal memproses token PPAA '{tokens_a[idx]}': {str(e)}")
                idx += 1
        except Exception as e: logs.append(f"Gagal membaca blok PPAA secara keseluruhan: {str(e)}")

    ppbb_data = {}
    if 'PPBB' in text:
        try:
            ppbb_part = text.split('PPBB', 1)[1]
            for end_mark in ['PPAA', 'PPCC', 'PPDD']:
                if end_mark in ppbb_part: ppbb_part = ppbb_part.split(end_mark)[0]
            tokens = ppbb_part.split()
            idx = 0
            while idx < len(tokens):
                try:
                    token = tokens[idx]
                    is_station_id = (idx < 2 and len(token) == 5 and token.startswith('97'))
                    if len(token) == 5 and token.startswith('9') and not is_station_id:
                        t_n = int(token[1]) if token[1].isdigit() else 0
                        levels = [int(ch) if ch.isdigit() else (0 if ch == '/' else None) for ch in token[2:5]]
                        valid_levels = [u for u in levels if u is not None]
                        for i, u in enumerate(levels):
                            try:
                                if u is not None and (idx + 1 + i) < len(tokens):
                                    wind_str = tokens[idx + 1 + i]
                                    if len(wind_str) == 5 and wind_str != '/////':
                                        lvl = (t_n * 10000) + (u * 1000)
                                        if lvl > 0: ppbb_data[f'{lvl} feet'] = (int(wind_str[:3]), int(wind_str[3:]))
                            except Exception as e: logs.append(f"Gagal ekstrak angin PPBB di sekitar token {token}: {str(e)}")
                        idx += len(valid_levels)
                except Exception as e: pass
                idx += 1
        except Exception as e: logs.append(f"Gagal membaca blok PPBB secara keseluruhan: {str(e)}")

    final_data = dict(ppaa_data)
    final_data.update(ppbb_data)
    final_data['PERMUKAAN'] = (surface_dir, surface_speed)
    return final_data, logs

def get_max_height_from_parsed(parsed_data):
    max_feet = 0
    for key in parsed_data.keys():
        if 'feet' in key:
            try:
                val = int(key.split(' ')[0])
                if val > max_feet: max_feet = val
            except ValueError: pass
    return max_feet if max_feet > 0 else 0

def process_template_uw1(wb, df_records, nama_bulan, sample_year, sample_month_num):
    ws_uw1 = wb['UW I'] if 'UW I' in wb.sheetnames else None
    ws_uw2 = wb['UW II'] if 'UW II' in wb.sheetnames else None
        
    col_mapping_uw1 = {
        '1000 feet': (12, 13, 14, 15), '2000 feet': (16, 17, 18, 19), 
        '3000 feet': (20, 21, 22, 23), '5000 feet': (24, 25, 26, 27), 
        '7000 feet': (28, 29, 30, 31), '10000 feet': (32, 33, 34, 35), 
        '12000 feet': (36, 37, 38, 39), '15000 feet': (40, 41, 42, 43)
    }
    col_mapping_uw2 = {
        '18000 feet': (12, 13, 14, 15), '20000 feet': (16, 17, 18, 19), 
        '25000 feet': (20, 21, 22, 23), '30000 feet': (24, 25, 26, 27), 
        '35000 feet': (28, 29, 30, 31), '40000 feet': (32, 33, 34, 35),
        '45000 feet': (36, 37, 38, 39), '50000 feet': (40, 41, 42, 43)
    }

    def get_row_uw(d, h):
        if h == 0: return 9 + (d - 1) * 3
        elif h == 6: return 9 + (d - 1) * 3 + 1
        elif h == 12: return 9 + (d - 1) * 3 + 2
        return None

    def split_digits(d_val, f_val):
        d_str = str(int(d_val)).zfill(3)[:2]
        f_str = str(int(f_val)).zfill(2)[-2:]
        return str(d_str[0]), str(d_str[1]), str(f_str[0]), str(f_str[1])
        
    year_2d = str(sample_year % 100).zfill(2)
    month_str = str(sample_month_num)

    for ws in [ws_uw1, ws_uw2]:
        if ws is None: continue
        safe_set_cell(ws, 4, 6, f": {nama_bulan}")
        safe_set_cell(ws, 5, 6, f": {sample_year}")
        row_limit = 9 if ws.title == 'UW I' else 8 
        for c in range(5, 10): safe_set_cell(ws, row_limit, c, None)
        safe_set_cell(ws, 9, 1, "9"); safe_set_cell(ws, 9, 2, "7"); safe_set_cell(ws, 9, 3, "3")
        safe_set_cell(ws, 9, 4, "4"); safe_set_cell(ws, 9, 5, "0"); safe_set_cell(ws, 9, 6, year_2d[0])
        safe_set_cell(ws, 9, 7, year_2d[1]); safe_set_cell(ws, 9, 8, month_str)
        safe_set_cell(ws, 9, 9, "0"); safe_set_cell(ws, 9, 10, "1")
            
    for _, row in df_records.iterrows():
        day = row['day']; hour = row['hour_z']
        target_row = get_row_uw(day, hour)
        if target_row is None: continue
        wind_dict = row['parsed_wind']
        hour_str = str(hour).zfill(2)
        
        if ws_uw1:
            safe_set_cell(ws_uw1, target_row, 11, hour_str) 
            for layer, cols in col_mapping_uw1.items():
                if layer in wind_dict:
                    d1, d2, f1, f2 = split_digits(*wind_dict[layer])
                    for i, val in enumerate([d1, d2, f1, f2]): safe_set_cell(ws_uw1, target_row, cols[i], val)
        if ws_uw2:
            safe_set_cell(ws_uw2, target_row, 11, hour_str) 
            for layer, cols in col_mapping_uw2.items():
                if layer in wind_dict:
                    d1, d2, f1, f2 = split_digits(*wind_dict[layer])
                    for i, val in enumerate([d1, d2, f1, f2]): safe_set_cell(ws_uw2, target_row, cols[i], val)
    return wb

def process_template_uw2(wb, df_records, nama_bulan, sample_year):
    for sheet_name in ['UPPER WIND', 'KETERANGAN']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            safe_set_cell(ws, 10, 6, f": {nama_bulan} {sample_year}")
            for r in range(16, 47):
                for c in range(3, 12): safe_set_cell(ws, r, c, None)
            for _, row in df_records.iterrows():
                day = row['day']; hour = row['hour_z']
                target_row = 15 + day
                col_pibal = 3 if hour == 0 else (6 if hour == 6 else (9 if hour == 12 else None))
                if col_pibal is not None:
                    max_height = row['max_height']
                    is_norep = (max_height == 0)
                    safe_set_cell(ws, target_row, col_pibal, "NOREP" if is_norep else f"{max_height:,}".replace(',', '.'))
                    alasan_final, pengamat_final = resolve_alasan_and_pengamat(row.get('alasan_data'), row.get('pengamat_data'), is_norep=is_norep)
                    safe_set_cell(ws, target_row, col_pibal + 1, alasan_final if alasan_final else "")
                    safe_set_cell(ws, target_row, col_pibal + 2, pengamat_final if pengamat_final else "-")
    return wb

def process_template_komponen(wb, df_records, nama_bulan, sample_year):
    START_ROW = 9
    sheet_mapping = {
        0:  {'P-5': '00 P-5', '7-20': '00 7-20', '25-40': '00 25-40', '45-50': '00 45-50'},
        6:  {'P-5': '06 P-5', '7-20': '06 7-20', '25-40': '06 25-40', '45-50': '06 45-50'},
        12: {'P-5': '12 P-5', '7-20': '12 7-20', '25-40': '12 25-40', '45-50': '12 45-50'},
    }
    col_mapping = {
        'PERMUKAAN': (2, 3), '1000 feet': (6, 7), '3000 feet': (10, 11), '5000 feet': (14, 15),
        '7000 feet': (2, 3), '10000 feet': (6, 7), '15000 feet': (10, 11), '20000 feet': (14, 15),
        '25000 feet': (2, 3), '30000 feet': (6, 7), '35000 feet': (10, 11), '40000 feet': (14, 15),
        '45000 feet': (2, 3), '50000 feet': (6, 7),
    }

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        safe_set_cell(ws, 3, 3, f": {nama_bulan} {sample_year}")
        for r in range(START_ROW, START_ROW + 31):
            for c in [2, 3, 6, 7, 10, 11, 14, 15]:
                if c <= ws.max_column: safe_set_cell(ws, r, c, None)

    layer_groups = [
        ('P-5', ['PERMUKAAN', '1000 feet', '3000 feet', '5000 feet']),
        ('7-20', ['7000 feet', '10000 feet', '15000 feet', '20000 feet']),
        ('25-40', ['25000 feet', '30000 feet', '35000 feet', '40000 feet']),
        ('45-50', ['45000 feet', '50000 feet'])
    ]

    for _, row in df_records.iterrows():
        hour, day = row['hour_z'], row['day']
        if hour not in sheet_mapping: continue
        target_row = START_ROW + (day - 1)
        wind_dict = row['parsed_wind']
        for group_key, group_layers in layer_groups:
            sheet_title = sheet_mapping[hour].get(group_key)
            if sheet_title and sheet_title in wb.sheetnames:
                ws_target = wb[sheet_title]
                for layer in group_layers:
                    if layer in wind_dict:
                        d, f = wind_dict[layer]
                        safe_set_cell(ws_target, target_row, col_mapping[layer][0], int(d))
                        safe_set_cell(ws_target, target_row, col_mapping[layer][1], int(f))
    return wb

# =========================================================
# ===== FORM A/B HUJAN HELLMAN & PENGUAPAN FUNCTIONS ======
# =========================================================

def get_val_special(val):
    """Fungsi helper khusus merubah 8888 -> TTU dan 9999 -> - """
    try:
        if pd.isna(val): return 0
        v = float(val)
        if v == 8888: return "TTU"
        if v == 9999: return "-"
        if v == int(v): return int(v)
        return v
    except:
        return 0

def process_form_ab(wb, df_hellman, df_penguapan, nama_bulan, sample_year, num_days):
    ws_a = wb['A'] if 'A' in wb.sheetnames else None
    ws_b1 = wb['B 1'] if 'B 1' in wb.sheetnames else None
    ws_b2 = wb['B 2'] if 'B 2' in wb.sheetnames else None

    if ws_a:
        safe_set_cell(ws_a, 5, 19, nama_bulan) # S5
        safe_set_cell(ws_a, 6, 19, sample_year) # S6
        
        col_map_a_intensitas = {
            'hellman_5mnt': 1, 'hellman_10mnt': 2, 'hellman_15mnt': 3,
            'hellman_30mnt': 4, 'hellman_45mnt': 5, 'hellman_60mnt': 6,
            'hellman_120mnt': 7, 'hellman_3jam': 8, 'hellman_6jam': 9,
            'hellman_12jam': 10
        }
        
        col_map_a_jam = {
            'hellman_07_08': 12, 'hellman_08_09': 13, 'hellman_09_10': 14, 'hellman_10_11': 15,
            'hellman_11_12': 16, 'hellman_12_13': 17, 'hellman_13_14': 18, 'hellman_14_15': 19,
            'hellman_15_16': 20, 'hellman_16_17': 21, 'hellman_17_18': 22, 'hellman_18_19': 23,
            'hellman_19_20': 24, 'hellman_20_21': 25, 'hellman_21_22': 26, 'hellman_22_23': 27,
            'hellman_23_24': 28, 'hellman_24_01': 29, 'hellman_01_02': 30, 'hellman_02_03': 31,
            'hellman_03_04': 32, 'hellman_04_05': 33, 'hellman_05_06': 34, 'hellman_06_07': 35
        }

        for d in range(1, 32):
            r = 11 + d
            if d <= num_days:
                safe_set_cell(ws_a, r, 11, d)
                for c in range(1, 36):
                    if c == 11: continue
                    safe_set_cell(ws_a, r, c, None)
            else:
                safe_set_cell(ws_a, r, 11, "-")
                for c in range(1, 36):
                    if c == 11: continue
                    safe_set_cell(ws_a, r, c, "-")
                
        max_values_a = {k: {'val': -0.0001, 'date': "-"} for k in col_map_a_intensitas.keys()}

        if df_hellman is not None and not df_hellman.empty:
            for _, row in df_hellman.iterrows():
                day = row['day']
                if day > num_days: continue
                target_row = 11 + day 

                for col_name, col_idx in col_map_a_intensitas.items():
                    val = get_val_special(row.get(col_name, 0))
                    safe_set_cell(ws_a, target_row, col_idx, val if val != 0 else 0)
                    if isinstance(val, (int, float)):
                        if val > max_values_a[col_name]['val']:
                            max_values_a[col_name]['val'] = val
                            max_values_a[col_name]['date'] = day

                for col_name, col_idx in col_map_a_jam.items():
                    val = get_val_special(row.get(col_name, 0))
                    safe_set_cell(ws_a, target_row, col_idx, val if val != 0 else 0)
                    
        for col_name, col_idx in col_map_a_intensitas.items():
            date_val = max_values_a[col_name]['date']
            safe_set_cell(ws_a, 44, col_idx, date_val if date_val != "-" else "")

    if ws_b1:
        safe_set_cell(ws_b1, 5, 17, nama_bulan) # Q5
        safe_set_cell(ws_b1, 6, 17, sample_year) # Q6
        
        col_map_b1_jam = {
            'hellman_07_08': 2, 'hellman_08_09': 3, 'hellman_09_10': 4, 'hellman_10_11': 5,
            'hellman_11_12': 6, 'hellman_12_13': 7, 'hellman_13_14': 8, 'hellman_14_15': 9,
            'hellman_15_16': 10, 'hellman_16_17': 11, 'hellman_17_18': 12, 'hellman_18_19': 13,
            'hellman_19_20': 14, 'hellman_20_21': 15, 'hellman_21_22': 16, 'hellman_22_23': 17,
            'hellman_23_24': 18, 'hellman_24_01': 19, 'hellman_01_02': 20, 'hellman_02_03': 21,
            'hellman_03_04': 32, 'hellman_04_05': 33, 'hellman_05_06': 34, 'hellman_06_07': 35
        }
        
        for r in range(12, 43):
            d = r - 11
            for c in range(2, 26):
                if d <= num_days:
                    safe_set_cell(ws_b1, r, c, None)
                else:
                    safe_set_cell(ws_b1, r, c, "-")
                
        if df_hellman is not None and not df_hellman.empty:
            for _, row in df_hellman.iterrows():
                day = row['day']
                if day > num_days: continue
                target_row = 11 + day 
                for col_name, col_idx in col_map_b1_jam.items():
                    val = get_val_special(row.get(col_name, 0))
                    safe_set_cell(ws_b1, target_row, col_idx, val if val != 0 else 0)

    if ws_b2:
        col_map_b2_intensitas = {
            'hellman_5mnt': 5, 'hellman_10mnt': 6, 'hellman_15mnt': 7,
            'hellman_30mnt': 8, 'hellman_45mnt': 9, 'hellman_60mnt': 10,
            'hellman_120mnt': 11, 'hellman_6jam': 12, 'hellman_12jam': 13
        }
        
        for r in range(5, 36):
            d = r - 4
            if d <= num_days:
                safe_set_cell(ws_b2, r, 2, None)
            else:
                safe_set_cell(ws_b2, r, 2, "-")
                
            for c in range(5, 14):
                if d <= num_days:
                    safe_set_cell(ws_b2, r, c, None)
                else:
                    safe_set_cell(ws_b2, r, c, "-")

        if df_penguapan is not None and not df_penguapan.empty:
            for _, row in df_penguapan.iterrows():
                dt = row['datetime']
                target_dt = dt - pd.Timedelta(days=1)
                
                if target_dt.month == df_penguapan['target_month'].iloc[0] and target_dt.year == df_penguapan['target_year'].iloc[0]:
                    target_day = target_dt.day
                    val = get_val_special(row.get('rr_0700', 0))
                    safe_set_cell(ws_b2, 4 + target_day, 2, val if val != 0 else 0)

        max_values_b2 = {k: {'val': -0.0001, 'date': "-"} for k in col_map_b2_intensitas.keys()}
                
        if df_hellman is not None and not df_hellman.empty:
            for _, row in df_hellman.iterrows():
                day = row['day']
                if day > num_days: continue
                target_row = 4 + day 
                
                for col_name, col_idx in col_map_b2_intensitas.items():
                    val = get_val_special(row.get(col_name, 0))
                    safe_set_cell(ws_b2, target_row, col_idx, val if val != 0 else 0)
                    
                    if isinstance(val, (int, float)):
                        if val > max_values_b2[col_name]['val']:
                            max_values_b2[col_name]['val'] = val
                            max_values_b2[col_name]['date'] = day
                    
        for col_name, col_idx in col_map_b2_intensitas.items():
            date_val = max_values_b2[col_name]['date']
            safe_set_cell(ws_b2, 37, col_idx, date_val if date_val != "-" else "")

    return wb

def process_template_penguapan(wb, df_penguapan_full, target_year, target_month, num_days, nama_bulan):
    ws = wb['Penguapan'] if 'Penguapan' in wb.sheetnames else wb.active
    
    safe_set_cell(ws, 2, 7, f":  {target_year}")
    safe_set_cell(ws, 3, 7, f":  {nama_bulan}")
    
    for d in range(16, 32):
        if d > num_days:
            r = d - 5 
            safe_set_cell(ws, r, 7, "-")  
            safe_set_cell(ws, r, 8, "-")  
            safe_set_cell(ws, r, 15, "-") 
            safe_set_cell(ws, r, 16, "-") 
            
    for _, row in df_penguapan_full.iterrows():
        dt = row['datetime']
        target_dt = dt - pd.Timedelta(days=1)
        
        if target_dt.month == target_month and target_dt.year == target_year:
            target_day = target_dt.day
            
            diff = get_val_special(row.get('op_diff_baca_0700', 0))
            rr = get_val_special(row.get('rr_0700', 0))
            ws_avg = get_val_special(row.get('ws_avg_50cm_0700', 0))
            t_air = get_val_special(row.get('t_air_avg_0700', 0))
            
            if target_day <= 15:
                r = 10 + target_day
                safe_set_cell(ws, r, 2, diff)
                safe_set_cell(ws, r, 3, rr)
                safe_set_cell(ws, r, 12, ws_avg)
                safe_set_cell(ws, r, 13, t_air)
            else:
                r = target_day - 5
                safe_set_cell(ws, r, 7, diff)
                safe_set_cell(ws, r, 8, rr)
                safe_set_cell(ws, r, 15, ws_avg)
                safe_set_cell(ws, r, 16, t_air)
            
    return wb

# =========================================================
# ===== LAMA PENYINARAN MATAHARI (LPM) FUNCTIONS ==========
# =========================================================

def process_template_lpm(wb, df_lpm_month, year, month, num_days, nama_bulan, kepala_nama, kepala_nip):
    ws = wb['LPM'] if 'LPM' in wb.sheetnames else wb.active
    
    safe_set_cell(ws, 5, 4, f": {nama_bulan}") # Cell D5
    safe_set_cell(ws, 6, 4, f": {year}")       # Cell D6
    
    col_mapping = {
        'ss_6_7': 3,   'ss_7_8': 4,   'ss_8_9': 5,   'ss_9_10': 6,
        'ss_10_11': 7, 'ss_11_12': 8, 'ss_12_13': 9, 'ss_13_14': 10,
        'ss_14-15': 11,'ss_15_16': 12,'ss_16_17': 13,'ss_17_18': 14
    }
    
    month_data = {}
    if df_lpm_month is not None and not df_lpm_month.empty:
        for _, r in df_lpm_month.iterrows():
            month_data[r['day']] = r
            
    last_observer = ""

    for d in range(1, 32):
        r = 13 + d
        
        if d > num_days:
            for col_idx in range(3, 19):
                safe_set_cell(ws, r, col_idx, "-")
        else:
            if d in month_data:
                row_data = month_data[d]
                
                for col_name, col_idx in col_mapping.items():
                    val = row_data.get(col_name)
                    if pd.notna(val):
                        try:
                            safe_set_cell(ws, r, col_idx, float(val))
                        except:
                            safe_set_cell(ws, r, col_idx, str(val))
                    else:
                        safe_set_cell(ws, r, col_idx, "-")
                
                tot8 = row_data.get('ss_total_8')
                tot12 = row_data.get('ss_total_12')
                safe_set_cell(ws, r, 15, float(tot8) if pd.notna(tot8) else "-")
                safe_set_cell(ws, r, 17, float(tot12) if pd.notna(tot12) else "-")
                
                if pd.notna(row_data.get('observer_name')):
                    last_observer = str(row_data['observer_name'])
            else:
                for col_idx in range(3, 19):
                    safe_set_cell(ws, r, col_idx, "-")
                    
    tgl_sekarang = datetime.now()
    tgl_ttd = f"WAINGAPU, {num_days} {nama_bulan} {year}"
    safe_set_cell(ws, 48, 13, tgl_ttd) # Cell M48
    safe_set_cell(ws, 55, 2, kepala_nama) # Cell B55
    if kepala_nip:
        safe_set_cell(ws, 56, 2, f"NIP. {kepala_nip}") # Cell B56
        
    if last_observer:
        safe_set_cell(ws, 55, 13, last_observer.upper()) # Cell M55
        
    return wb

# ==========================================
# ======== ANTARMUKA WEB STREAMLIT =========
# ==========================================
st.sidebar.title("🎛️ Navigasi Menu")
menu = st.sidebar.radio(
    "Pilih Converter", 
    [
        "METAR Converter", 
        "SPECI Converter", 
        "WXREV Converter", 
        "Thunderstorm Exporter", 
        "Form Perawanan Exporter",
        "Form Komponen Angin, UWI, dan UWII",
        "Form A/B Penakar Hujan dan Penguapan",
        "Form Lama Penyinaran Matahari (LPM)"
    ]
)
st.sidebar.markdown("---")
st.sidebar.info("Aplikasi ekstraksi data Sandi Cuaca (METAR, SPECI, WXREV, Thunderstorm, Perawanan, Pibal, Hujan/Penguapan, & LPM) menjadi Laporan Excel otomatis.")

LOGO_FILE = "logo_bmkg.png"

if not os.path.exists(LOGO_FILE):
    st.sidebar.warning(f"⚠️ File gambar '{LOGO_FILE}' tidak terdeteksi di folder utama.")

# --- HALAMAN METAR ---
if menu == "METAR Converter":
    st.title("✈️ METAR to Excel Converter")
    with st.expander("ℹ️ Klik di sini untuk melihat Petunjuk Penggunaan"):
        st.markdown("""
        **Syarat File CSV:**
        - File hasil extract dari "https://bmkgsatu.bmkg.go.id/extractgts" data METAR dengan status SENT.
        - Jika file `TEMPLATE METAR_3.xlsx` tersedia di direktori yang sama, program akan menggunakan format KOP Surat terstandar.
        """)
    uploaded_file = st.file_uploader("Upload file CSV METAR", type=["csv"])
    if uploaded_file is not None:
        try:
            df = pd.read_csv(uploaded_file)
            if 'sandi' not in df.columns or 'data_timestamp' not in df.columns:
                st.error("Format CSV tidak sesuai! Pastikan terdapat kolom 'sandi' dan 'data_timestamp'.")
            else:
                with st.spinner("Sedang memvalidasi data METAR..."):
                    parsed_rows = []
                    for idx, row in df.iterrows():
                        res = parse_metar_speci(row['sandi'])
                        if res and res[0] == 'METAR':
                            station = row['station_name'] if 'station_name' in df.columns else "STASIUN METEOROLOGI"
                            msg_id = row['id'] if 'id' in df.columns else idx
                            parsed_rows.append(res + [row['data_timestamp'], station, msg_id])
                            
                    columns = ['TYPE', 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK', 'is_cor', 'cc_type', 'raw_timestamp', 'station_name', 'msg_id']
                    df_clean = pd.DataFrame(parsed_rows, columns=columns)
                    
                    if df_clean.empty: st.warning("Tidak ditemukan data METAR di dalam file ini.")
                    else:
                        df_clean['raw_timestamp'] = df_clean['raw_timestamp'].str.replace(" +0000 UTC", "", regex=False)
                        df_clean['datetime'] = pd.to_datetime(df_clean['raw_timestamp'])
                        df_clean = df_clean[df_clean['datetime'].dt.minute == 0]
                        df_clean['priority_score'] = df_clean.apply(calculate_priority, axis=1)
                        df_clean = df_clean.sort_values(by=['datetime', 'priority_score', 'msg_id'])
                        df_clean = df_clean.drop_duplicates(subset=['datetime'], keep='last')
                        df_clean['date_group'] = df_clean['datetime'].dt.date
                        df_clean = df_clean.sort_values(by='datetime').reset_index(drop=True)
                        
                        if df_clean.empty: st.warning("Tidak ditemukan data METAR dengan menit :00 (per jam) di dalam file ini.")
                        else:
                            st.success(f"Berhasil memproses data METAR!")
                            
                            DEFAULT_TEMPLATE_METAR = "TEMPLATE METAR_3.xlsx"
                            if os.path.exists(DEFAULT_TEMPLATE_METAR) or os.path.exists("TEMPLATE METAR_2.xlsx") or os.path.exists("TEMPLATE METAR.xlsx"):
                                excel_data = generate_excel_from_template_daily(df_clean, report_type="METAR", template_path=DEFAULT_TEMPLATE_METAR, logo_path=LOGO_FILE)
                            else:
                                st.warning(f"⚠️ File '{DEFAULT_TEMPLATE_METAR}' tidak ditemukan. Menggunakan format Excel standar.")
                                excel_data = generate_excel_bytes_metar_speci_fallback(df_clean, report_type="METAR")
                            
                            first_date = df_clean['date_group'].iloc[0]
                            nama_file_base = f"REKAP_METAR_{first_date.day:02d}_{BULAN_INDO[first_date.month]}_{first_date.year}"
                            
                            st.write("---")
                            st.subheader("Unduh Laporan METAR")
                            st.download_button(label="📊 Download Excel", data=excel_data, file_name=f"{nama_file_base}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e: st.error(f"Terjadi kesalahan: {e}")

# --- HALAMAN SPECI ---
elif menu == "SPECI Converter":
    st.title("🛩️ SPECI to Excel Converter")
    with st.expander("ℹ️ Klik di sini untuk melihat Petunjuk Penggunaan"):
        st.markdown("""
        **Syarat File CSV:**
        - File hasil extract dari "https://bmkgsatu.bmkg.go.id/extractgts" data SPECI dengan status SENT.
        """)
    uploaded_file = st.file_uploader("Upload file CSV SPECI", type=["csv"])
    if uploaded_file is not None:
        try:
            df = pd.read_csv(uploaded_file)
            if 'sandi' not in df.columns or 'data_timestamp' not in df.columns:
                st.error("Format CSV tidak sesuai! Pastikan terdapat kolom 'sandi' dan 'data_timestamp'.")
            else:
                with st.spinner("Sedang memvalidasi data SPECI..."):
                    parsed_rows = []
                    for idx, row in df.iterrows():
                        res = parse_metar_speci(row['sandi'])
                        if res and res[0] == 'SPECI':
                            station = row['station_name'] if 'station_name' in df.columns else "STASIUN METEOROLOGI"
                            msg_id = row['id'] if 'id' in df.columns else idx
                            parsed_rows.append(res + [row['data_timestamp'], station, msg_id])
                            
                    columns = ['TYPE', 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK', 'is_cor', 'cc_type', 'raw_timestamp', 'station_name', 'msg_id']
                    df_clean = pd.DataFrame(parsed_rows, columns=columns)
                    
                    if df_clean.empty: st.warning("Tidak ditemukan data SPECI di dalam file ini.")
                    else:
                        df_clean['raw_timestamp'] = df_clean['raw_timestamp'].str.replace(" +0000 UTC", "", regex=False)
                        df_clean['datetime'] = pd.to_datetime(df_clean['raw_timestamp'])
                        df_clean['priority_score'] = df_clean.apply(calculate_priority, axis=1)
                        df_clean = df_clean.sort_values(by=['datetime', 'priority_score', 'msg_id'])
                        df_clean = df_clean.drop_duplicates(subset=['datetime'], keep='last')
                        df_clean['date_group'] = df_clean['datetime'].dt.date
                        df_clean = df_clean.sort_values(by='datetime').reset_index(drop=True)
                        
                        st.success(f"Berhasil memproses {len(df_clean)} data SPECI!")
                        
                        DEFAULT_TEMPLATE_METAR = "TEMPLATE METAR_3.xlsx"
                        if os.path.exists(DEFAULT_TEMPLATE_METAR) or os.path.exists("TEMPLATE METAR_2.xlsx") or os.path.exists("TEMPLATE METAR.xlsx"):
                            excel_data = generate_excel_from_template_speci(df_clean, report_type="SPECI", template_path=DEFAULT_TEMPLATE_METAR, logo_path=LOGO_FILE)
                        else:
                            st.warning(f"⚠️ File '{DEFAULT_TEMPLATE_METAR}' tidak ditemukan. Menggunakan format Excel standar.")
                            excel_data = generate_excel_bytes_metar_speci_fallback(df_clean, report_type="SPECI")
                        
                        first_date = df_clean['datetime'].iloc[0]
                        nama_file_base = f"REKAP_SPECI_BULAN_{BULAN_INDO[first_date.month]}_{first_date.year}"
                        
                        st.write("---")
                        st.subheader("Unduh Laporan SPECI")
                        st.download_button(label="📊 Download Excel", data=excel_data, file_name=f"{nama_file_base}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e: st.error(f"Terjadi kesalahan: {e}")

# --- HALAMAN WXREV ---
elif menu == "WXREV Converter":
    st.title("🌧️ WXREV to Excel Converter")
    with st.expander("ℹ️ Klik di sini untuk melihat Petunjuk Penggunaan"):
        st.markdown("""
        **Syarat File CSV:**
        - File hasil extract dari "https://bmkgsatu.bmkg.go.id/extractgts" data WXREV.
        - Menggunakan template `TEMPLATE WXREV.xlsx`.
        """)
    uploaded_file = st.file_uploader("Upload file CSV WXREV", type=["csv"])
    if uploaded_file is not None:
        try:
            df = pd.read_csv(uploaded_file)
            if 'sandi' not in df.columns or 'data_timestamp' not in df.columns:
                st.error("Format CSV tidak sesuai! Pastikan terdapat kolom 'sandi' dan 'data_timestamp'.")
            else:
                with st.spinner("Sedang memvalidasi data WXREV..."):
                    parsed_rows = []
                    for idx, row in df.iterrows():
                        res = parse_wxrev(row['sandi'])
                        if res:
                            station = row['station_name'] if 'station_name' in df.columns else "Stasiun Meteorologi"
                            parsed_rows.append(res + [row['data_timestamp'], station])
                            
                    if not parsed_rows: st.warning("Tidak ditemukan data sandi WXREV di dalam file ini.")
                    else:
                        columns = ['TGL', 'MMYYGp', 'IIiii', 'atTxTxTnTn', 'apPxPxPnPn', 'auUxUxUnUn', 'arRRRR', 'rDrDdfmfm_1', 'rDrDdfmfm_2', 'raw_timestamp', 'station_name']
                        df_wxrev = pd.DataFrame(parsed_rows, columns=columns)
                        df_wxrev['raw_timestamp'] = df_wxrev['raw_timestamp'].str.replace(" +0000 UTC", "", regex=False)
                        df_wxrev['datetime'] = pd.to_datetime(df_wxrev['raw_timestamp'])
                        df_wxrev = df_wxrev.sort_values(by='datetime')
                        df_wxrev = df_wxrev.drop_duplicates(subset=['TGL'], keep='last')
                        df_wxrev = df_wxrev.sort_values(by='TGL').reset_index(drop=True)
                        
                        st.success(f"Berhasil memproses {len(df_wxrev)} data WXREV harian!")
                        
                        DEFAULT_TEMPLATE_WXREV = "TEMPLATE WXREV.xlsx"
                        if os.path.exists(DEFAULT_TEMPLATE_WXREV):
                            excel_data = generate_excel_from_template_wxrev(df_wxrev, template_path=DEFAULT_TEMPLATE_WXREV)
                        else:
                            st.warning(f"⚠️ File '{DEFAULT_TEMPLATE_WXREV}' tidak ditemukan. Menggunakan format Excel standar.")
                            excel_data = generate_excel_bytes_wxrev(df_wxrev)
                            
                        dt_first = df_wxrev['datetime'].iloc[0]
                        nama_file_base = f"REKAP_WXREV_{BULAN_INDO.get(dt_first.month, '').upper()}_{dt_first.year}"
                        
                        st.write("---")
                        st.subheader("Unduh Laporan WXREV")
                        st.download_button(label="📊 Download Excel", data=excel_data, file_name=f"{nama_file_base}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e: st.error(f"Terjadi kesalahan: {e}")

# --- HALAMAN THUNDERSTORM EXPORTER ---
elif menu == "Thunderstorm Exporter":
    st.title("⚡ Thunderstorm Data Exporter")
    st.write("Ekstraksi data kejadian Kilat / Thunderstorm tahunan dari gabungan file METAR & SPECI.")
    with st.expander("⚙️ Pengaturan Header & Tanda Tangan PDF"):
        st_nama = st.text_input("Nama Stasiun", "STASIUN METEOROLOGI UMBU MEHANG KUNDA")
        kp_nama = st.text_input("Nama Kepala Stasiun", "CARLES ALEXANDER TARI, S.TP")
        kp_nip = st.text_input("NIP Kepala Stasiun (Opsional)", "")
    uploaded_files = st.file_uploader("Upload File CSV (Bisa Upload Multiple METAR/SPECI)", type=["csv"], accept_multiple_files=True)
    if uploaded_files:
        all_rows = []
        for file in uploaded_files:
            try:
                df = pd.read_csv(file)
                if 'sandi' in df.columns and 'data_timestamp' in df.columns:
                    for idx, row in df.iterrows():
                        res = parse_metar_speci(row['sandi'])
                        if res:
                            station = row['station_name'] if 'station_name' in df.columns else st_nama
                            all_rows.append(res + [row['data_timestamp'], station])
            except Exception as e: st.error(f"Gagal membaca file {file.name}: {e}")
        if all_rows:
            columns = ['TYPE', 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK', 'is_cor', 'cc_type', 'raw_timestamp', 'station_name']
            df_clean = pd.DataFrame(all_rows, columns=columns)
            df_clean['raw_timestamp'] = df_clean['raw_timestamp'].str.replace(" +0000 UTC", "", regex=False)
            df_clean['datetime'] = pd.to_datetime(df_clean['raw_timestamp'])
            df_clean = df_clean.sort_values(by='datetime').reset_index(drop=True)
            pdf_data = generate_pdf_bytes_thunderstorm(df_clean, st_nama, kp_nama, kp_nip)
            excel_data = generate_excel_bytes_thunderstorm(df_clean, st_nama)
            tahun = df_clean['datetime'].dt.year.iloc[0]
            st.success(f"Berhasil memproses data Thunderstorm Tahun {tahun}!")
            st.write("---")
            col_pdf, col_xlsx = st.columns(2)
            with col_pdf: st.download_button(label="📥 Download PDF Thunderstorm", data=pdf_data, file_name=f"THUNDERSTORM_{tahun}.pdf", mime="application/pdf")
            with col_xlsx: st.download_button(label="📊 Download Excel Thunderstorm", data=excel_data, file_name=f"THUNDERSTORM_{tahun}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- HALAMAN FORM PERAWANAN EXPORTER ---
elif menu == "Form Perawanan Exporter":
    st.title("☁️ Form Perawanan Exporter")
    st.write("Ekstraksi data Ikhtisar Frekuensi Perawanan bulanan dari file METAR per jam.")
    with st.expander("⚙️ Pengaturan Header, Form & Tanda Tangan PDF"):
        st_nama = st.text_input("Nama Stasiun", "STASIUN METEOROLOGI UMBU MEHANG KUNDA", key="pwn_st")
        form_code = st.text_input("Kode Form (Pojok Kiri Bawah)", "Klim / PWN/WPU-2001")
        kp_nama = st.text_input("Nama Kepala Stasiun", "CARLES ALEXANDER TARI, S.TP", key="pwn_kp")
        kp_nip = st.text_input("NIP Kepala Stasiun", "197712082001121001", key="pwn_nip")
    uploaded_files = st.file_uploader("Upload File CSV METAR", type=["csv"], accept_multiple_files=True, key="pwn_uploader")
    if uploaded_files:
        all_rows = []
        for file in uploaded_files:
            try:
                df = pd.read_csv(file)
                if 'sandi' in df.columns and 'data_timestamp' in df.columns:
                    for idx, row in df.iterrows():
                        res = parse_metar_speci(row['sandi'])
                        if res and res[0] == 'METAR':
                            station = row['station_name'] if 'station_name' in df.columns else st_nama
                            msg_id = row['id'] if 'id' in df.columns else idx
                            all_rows.append(res + [row['data_timestamp'], station, msg_id])
            except Exception as e: st.error(f"Gagal membaca file {file.name}: {e}")
        if all_rows:
            columns = ['TYPE', 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK', 'is_cor', 'cc_type', 'raw_timestamp', 'station_name', 'msg_id']
            df_clean = pd.DataFrame(all_rows, columns=columns)
            df_clean['raw_timestamp'] = df_clean['raw_timestamp'].str.replace(" +0000 UTC", "", regex=False)
            df_clean['datetime'] = pd.to_datetime(df_clean['raw_timestamp'])
            df_clean = df_clean[df_clean['datetime'].dt.minute == 0]
            df_clean['priority_score'] = df_clean.apply(calculate_priority, axis=1)
            df_clean = df_clean.sort_values(by=['datetime', 'priority_score', 'msg_id'])
            df_clean = df_clean.drop_duplicates(subset=['datetime'], keep='last')
            df_clean = df_clean.sort_values(by='datetime').reset_index(drop=True)
            if df_clean.empty: st.warning("Tidak ditemukan data METAR dengan menit :00 di dalam file ini.")
            else:
                pdf_data = generate_pdf_bytes_perawanan(df_clean, st_nama, kp_nama, kp_nip, form_code)
                excel_data = generate_excel_bytes_perawanan(df_clean, st_nama)
                months_count = len(df_clean.groupby([df_clean['datetime'].dt.year, df_clean['datetime'].dt.month]))
                dt_min = df_clean['datetime'].min(); dt_max = df_clean['datetime'].max()
                st.success(f"Berhasil memproses Data Perawanan untuk {months_count} bulan ({dt_min.strftime('%b %Y')} - {dt_max.strftime('%b %Y')})!")
                st.write("---")
                col_pdf, col_xlsx = st.columns(2)
                with col_pdf: st.download_button(label="📥 Download PDF Perawanan", data=pdf_data, file_name=f"PERAWANAN_{dt_min.strftime('%Y%m')}_{dt_max.strftime('%Y%m')}.pdf", mime="application/pdf")
                with col_xlsx: st.download_button(label="📊 Download Excel Perawanan", data=excel_data, file_name=f"PERAWANAN_{dt_min.strftime('%Y%m')}_{dt_max.strftime('%Y%m')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- HALAMAN PIBAL / RADIOSONDE ---
elif menu == "Form Komponen Angin, UWI, dan UWII":
    st.title("🎈 Otomatisasi Rekap Udara Atas (Pibal / Radiosonde)")
    st.subheader("Stasiun Meteorologi BMKG")
    st.markdown("---")

    DEFAULT_TEMPLATE_UW1 = "template uw 1.xlsx"
    DEFAULT_TEMPLATE_UW2 = "template uw 2.xlsx"
    DEFAULT_TEMPLATE_KOMPONEN = "template komponen angin.xlsx"

    uploaded_pibal_files = st.file_uploader(
        "Upload File CSV BMKG (Bisa Upload 2 File Sekaligus: Raw Pibal & Data Pibal)", 
        type=['csv'], 
        accept_multiple_files=True
    )

    if uploaded_pibal_files:
        use_uw1 = DEFAULT_TEMPLATE_UW1 if os.path.exists(DEFAULT_TEMPLATE_UW1) else None
        use_uw2 = DEFAULT_TEMPLATE_UW2 if os.path.exists(DEFAULT_TEMPLATE_UW2) else None
        use_komp = DEFAULT_TEMPLATE_KOMPONEN if os.path.exists(DEFAULT_TEMPLATE_KOMPONEN) else None

        if not use_uw1 and not use_uw2 and not use_komp:
            st.error("❌ File Template Excel lokal tidak ditemukan di direktori aplikasi! Pastikan file template UW1, UW2, dan Komponen Angin ada di folder utama bersama skrip.")
        else:
            try:
                df_raw = None; df_data = None
                for file in uploaded_pibal_files:
                    df_temp = pd.read_csv(file)
                    if 'data_timestamp' not in df_temp.columns:
                        file.seek(0)
                        df_temp = pd.read_csv(file, skiprows=1)
                    cols = [c.lower() for c in df_temp.columns]
                    if 'lapisan' in cols or 'wd' in cols: df_data = df_temp
                    elif 'pembacaan' in cols or 'observer_name' in cols or 'sandi_pibal' in cols: df_raw = df_temp
                    else:
                        if df_raw is None: df_raw = df_temp

                if df_raw is None and df_data is not None: df_raw = df_data

                col_alasan, col_pengamat = detect_columns(df_raw.columns)
                timestamps = df_raw['data_timestamp'].unique()
                records = []

                for ts in timestamps:
                    sub_raw = df_raw[df_raw['data_timestamp'] == ts]
                    raw_alasan = extract_valid_value(sub_raw[col_alasan]) if col_alasan and col_alasan in sub_raw.columns else None
                    raw_pengamat = extract_valid_value(sub_raw[col_pengamat]) if col_pengamat and col_pengamat in sub_raw.columns else None
                    s_dir = sub_raw['wind_dir_surface'].iloc[0] if 'wind_dir_surface' in sub_raw.columns else 0
                    s_spd = sub_raw['wind_speed_surface'].iloc[0] if 'wind_speed_surface' in sub_raw.columns else 0
                    sandi = extract_valid_value(sub_raw['sandi_pibal']) if 'sandi_pibal' in sub_raw.columns else ""

                    wind_dict = {}
                    parse_logs = []
                    if df_data is not None and 'lapisan' in df_data.columns:
                        sub_data = df_data[df_data['data_timestamp'] == ts]
                        wind_dict['PERMUKAAN'] = (s_dir, s_spd)
                        for _, d_row in sub_data.iterrows():
                            if pd.notna(d_row.get('lapisan')) and pd.notna(d_row.get('wd')) and pd.notna(d_row.get('ws')):
                                wind_dict[f"{int(d_row['lapisan'])} feet"] = (int(d_row['wd']), int(d_row['ws']))
                    
                    if not wind_dict or len(wind_dict) <= 1:
                        wind_dict, parse_logs = parse_sandi_pibal_complete(sandi, s_dir, s_spd)

                    max_h = get_max_height_from_parsed(wind_dict)
                    dt = pd.to_datetime(ts)
                    records.append({
                        'data_timestamp': ts, 'day': dt.day, 'month': dt.month, 'year': dt.year, 'hour_z': dt.hour,
                        'alasan_data': raw_alasan, 'pengamat_data': raw_pengamat, 'parsed_wind': wind_dict,
                        'max_height': max_h, 'parse_logs': parse_logs
                    })

                df_records = pd.DataFrame(records).sort_values(['month', 'day', 'hour_z']).reset_index(drop=True)
                error_records = [r for r in records if len(r.get('parse_logs', [])) > 0]
                if error_records:
                    with st.expander("⚠️ Peringatan: Terdapat Isu Parsing Sandi Pibal (Klik untuk detail)"):
                        st.warning(f"Ditemukan {len(error_records)} pengamatan dengan sandi malformed/korup. Sistem mengabaikan blok yang rusak dan mengambil data yang valid.")
                        for err in error_records:
                            st.markdown(f"**Waktu Pengamatan (UTC):** `{err['data_timestamp']}`")
                            for log_msg in err['parse_logs']: st.markdown(f" - {log_msg}")

                sample_month_num = df_records['month'].iloc[0]
                sample_year = df_records['year'].iloc[0]
                nama_bulan = BULAN_INDO[sample_month_num]

                st.success(f"✅ Data berhasil digabungkan: **{len(df_records)} pengamatan** periode **{nama_bulan} {sample_year}**.")
                if df_data is not None: st.info("💡 **Mode Dual CSV Aktif**: Data angin per lapisan diambil matang dari `Data Pibal`. Metadata diambil dari `Raw Pibal`.")
                else: st.info("ℹ️ **Mode Single CSV**: Data ditarik via Parser Sandi Pibal WMO.")

                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                    if use_uw1:
                        wb_uw1 = openpyxl.load_workbook(use_uw1)
                        wb_uw1 = process_template_uw1(wb_uw1, df_records, nama_bulan, sample_year, sample_month_num)
                        excel_buffer_uw1 = io.BytesIO(); wb_uw1.save(excel_buffer_uw1); excel_buffer_uw1.seek(0)
                        zip_file.writestr(f"Laporan_UW1_{nama_bulan}_{sample_year}.xlsx", excel_buffer_uw1.getvalue())
                        st.info("📋 Template UW 1 (UW I & UW II) berhasil diproses.")
                    if use_uw2:
                        wb_uw2 = openpyxl.load_workbook(use_uw2)
                        wb_uw2 = process_template_uw2(wb_uw2, df_records, nama_bulan, sample_year)
                        excel_buffer_uw2 = io.BytesIO(); wb_uw2.save(excel_buffer_uw2); excel_buffer_uw2.seek(0)
                        zip_file.writestr(f"Laporan_UW2_{nama_bulan}_{sample_year}.xlsx", excel_buffer_uw2.getvalue())
                        st.info("📋 Template UW 2 (Form Ae.31) berhasil diproses.")
                    if use_komp:
                        wb_komp = openpyxl.load_workbook(use_komp)
                        wb_komp = process_template_komponen(wb_komp, df_records, nama_bulan, sample_year)
                        excel_buffer_komp = io.BytesIO(); wb_komp.save(excel_buffer_komp); excel_buffer_komp.seek(0)
                        zip_file.writestr(f"Rekap_Komponen_Angin_{nama_bulan}_{sample_year}.xlsx", excel_buffer_komp.getvalue())
                        st.info("📋 Template Rekap Komponen Angin (Lengkap s.d 50.000 ft) berhasil diproses.")

                st.subheader("📥 Unduh Hasil Rekapitulasi")
                zip_buffer.seek(0)
                st.download_button(
                    label="⬇️ Download Semua File (Bentuk ZIP)",
                    data=zip_buffer,
                    file_name=f"Rekap_Udara_Atas_{nama_bulan}_{sample_year}.zip",
                    mime="application/zip",
                )
            except Exception as e:
                st.error(f"❌ Terjadi kesalahan saat memproses data: {e}")

# --- HALAMAN FORM A/B PENAKAR HUJAN & PENGUAPAN ---
elif menu == "Form A/B Penakar Hujan dan Penguapan":
    st.title("🌧️ Form A/B Penakar Hujan dan Penguapan")
    st.write("Ekstraksi data CSV Hujan Hellman & OP Penguapan ke Template Excel secara otomatis.")
    
    DEFAULT_TEMPLATE_AB = "TEMPLATE FORM A B.xlsx"
    DEFAULT_TEMPLATE_PENGUAPAN = "TEMPLATE PENGUAPAN.xlsx"
    
    uploaded_files = st.file_uploader("Upload File CSV (Bisa Hujan Hellman saja, Penguapan saja, atau Keduanya)", type=["csv"], accept_multiple_files=True)
    
    if uploaded_files:
        if not os.path.exists(DEFAULT_TEMPLATE_AB) or not os.path.exists(DEFAULT_TEMPLATE_PENGUAPAN):
            st.error(f"❌ File '{DEFAULT_TEMPLATE_AB}' atau '{DEFAULT_TEMPLATE_PENGUAPAN}' tidak ditemukan di direktori aplikasi!")
        else:
            try:
                df_hellman = pd.DataFrame()
                df_penguapan = pd.DataFrame()
                
                for file in uploaded_files:
                    df_temp = pd.read_csv(file)
                    if df_temp.shape[1] == 1: 
                        file.seek(0)
                        df_temp = pd.read_csv(file, skiprows=1)
                        
                    if 'hellman_5mnt' in df_temp.columns:
                        df_hellman = df_temp
                    elif 'op_eva_0700' in df_temp.columns:
                        df_penguapan = df_temp
                
                if df_hellman.empty and df_penguapan.empty:
                    st.error("Kolom data yang diperlukan tidak ditemukan pada file CSV yang diunggah.")
                else:
                    with st.spinner("Memproses data..."):
                        all_ym = set()
                        
                        if not df_hellman.empty:
                            df_hellman['datetime'] = pd.to_datetime(df_hellman['data_timestamp'])
                            df_hellman['day'] = df_hellman['datetime'].dt.day
                            df_hellman['month'] = df_hellman['datetime'].dt.month
                            df_hellman['year'] = df_hellman['datetime'].dt.year
                            all_ym.update(zip(df_hellman['year'], df_hellman['month']))
                            
                        if not df_penguapan.empty:
                            df_penguapan['datetime'] = pd.to_datetime(df_penguapan['data_timestamp'])
                            df_penguapan['day'] = df_penguapan['datetime'].dt.day
                            df_penguapan['month'] = df_penguapan['datetime'].dt.month
                            df_penguapan['year'] = df_penguapan['datetime'].dt.year
                            
                            df_penguapan['target_date'] = df_penguapan['datetime'] - pd.Timedelta(days=1)
                            df_penguapan['target_month'] = df_penguapan['target_date'].dt.month
                            df_penguapan['target_year'] = df_penguapan['target_date'].dt.year
                            
                            all_ym.update(zip(df_penguapan['target_year'], df_penguapan['target_month']))
                        
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                            file_count = 0
                            
                            for year, month in sorted(list(all_ym)):
                                nama_bulan = BULAN_INDO[month]
                                num_days = calendar.monthrange(year, month)[1]
                                
                                df_h_month = pd.DataFrame()
                                if not df_hellman.empty:
                                    df_h_month = df_hellman[(df_hellman['year'] == year) & (df_hellman['month'] == month)].sort_values('day').reset_index(drop=True)
                                
                                df_p_month = pd.DataFrame()
                                if not df_penguapan.empty:
                                    df_p_month = df_penguapan[
                                        (df_penguapan['target_year'] == year) & (df_penguapan['target_month'] == month)
                                    ].sort_values('datetime').reset_index(drop=True)
                                
                                if not df_h_month.empty or not df_p_month.empty:
                                    wb_ab = openpyxl.load_workbook(DEFAULT_TEMPLATE_AB)
                                    wb_ab = process_form_ab(wb_ab, df_h_month, df_p_month, nama_bulan, year, num_days)
                                    excel_buffer_ab = io.BytesIO()
                                    wb_ab.save(excel_buffer_ab)
                                    excel_buffer_ab.seek(0)
                                    zip_file.writestr(f"Form_AB_{nama_bulan}_{year}.xlsx", excel_buffer_ab.getvalue())
                                
                                if not df_p_month.empty:
                                    wb_peng = openpyxl.load_workbook(DEFAULT_TEMPLATE_PENGUAPAN)
                                    wb_peng = process_template_penguapan(wb_peng, df_p_month, year, month, num_days, nama_bulan)
                                    excel_buffer_peng = io.BytesIO()
                                    wb_peng.save(excel_buffer_peng)
                                    excel_buffer_peng.seek(0)
                                    zip_file.writestr(f"Form_Penguapan_{nama_bulan}_{year}.xlsx", excel_buffer_peng.getvalue())
                                    
                                file_count += 1
                                
                        st.success(f"✅ Berhasil memproses data untuk {file_count} siklus bulanan.")
                        st.info("💡 Semua data yang telah diproses digabungkan menjadi file zip yang rapih di bawah.")
                        
                        zip_buffer.seek(0)
                        st.download_button(
                            label="⬇️ Download Hasil Form (Bentuk ZIP)",
                            data=zip_buffer,
                            file_name="Rekap_Hujan_Penguapan_Otomatis.zip",
                            mime="application/zip",
                        )
            except Exception as e:
                st.error(f"❌ Terjadi kesalahan saat memproses data: {e}")

# --- HALAMAN FORM LAMA PENYINARAN MATAHARI (LPM) ---
elif menu == "Form Lama Penyinaran Matahari (LPM)":
    st.title("☀️ Form Lama Penyinaran Matahari (LPM)")
    st.write("Ekstraksi data CSV Penyinaran Matahari ke Template Excel (TEMPLATE LPM.xlsx) secara otomatis.")
    
    DEFAULT_TEMPLATE_LPM = "TEMPLATE LPM.xlsx"
    
    with st.expander("⚙️ Pengaturan Header & Tanda Tangan"):
        kp_nama = st.text_input("Nama Kepala Stasiun", "Carles Alexander Tari, S.TP", key="lpm_kp")
        kp_nip = st.text_input("NIP Kepala Stasiun", "197712082001121001", key="lpm_nip")
        
    uploaded_lpm_file = st.file_uploader("Upload File CSV Lama Penyinaran Matahari", type=["csv"], key="lpm_uploader")
    
    if uploaded_lpm_file:
        if not os.path.exists(DEFAULT_TEMPLATE_LPM):
            st.error(f"❌ File '{DEFAULT_TEMPLATE_LPM}' tidak ditemukan di direktori aplikasi!")
        else:
            try:
                # Membaca file CSV
                df_lpm = pd.read_csv(uploaded_lpm_file)
                if df_lpm.shape[1] == 1:
                    uploaded_lpm_file.seek(0)
                    df_lpm = pd.read_csv(uploaded_lpm_file, skiprows=1)
                    
                if 'data_timestamp' not in df_lpm.columns or 'ss_6_7' not in df_lpm.columns:
                    st.error("Kolom data yang diperlukan ('data_timestamp', 'ss_6_7', dll.) tidak ditemukan pada file CSV yang diunggah.")
                else:
                    with st.spinner("Memproses data Penyinaran Matahari..."):
                        df_lpm['datetime'] = pd.to_datetime(df_lpm['data_timestamp'])
                        df_lpm['day'] = df_lpm['datetime'].dt.day
                        df_lpm['month'] = df_lpm['datetime'].dt.month
                        df_lpm['year'] = df_lpm['datetime'].dt.year
                        
                        grouped_months = df_lpm.groupby(['year', 'month'])
                        
                        processed_files = []
                        
                        for (year, month), month_group in grouped_months:
                            nama_bulan = BULAN_INDO[month]
                            num_days = calendar.monthrange(year, month)[1]
                            
                            wb_lpm = openpyxl.load_workbook(DEFAULT_TEMPLATE_LPM)
                            wb_lpm = process_template_lpm(wb_lpm, month_group, year, month, num_days, nama_bulan, kp_nama, kp_nip)
                            
                            excel_buffer = io.BytesIO()
                            wb_lpm.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            file_name = f"LPM_{nama_bulan}_{year}.xlsx"
                            processed_files.append((file_name, excel_buffer.getvalue()))
                            
                        st.success(f"✅ Berhasil memproses data untuk {len(processed_files)} bulan.")
                        
                        if len(processed_files) == 1:
                            file_name, file_bytes = processed_files[0]
                            st.download_button(
                                label="📊 Download File Excel LPM",
                                data=file_bytes,
                                file_name=file_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            zip_buffer = io.BytesIO()
                            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                                for file_name, file_bytes in processed_files:
                                    zip_file.writestr(file_name, file_bytes)
                                    
                            zip_buffer.seek(0)
                            st.download_button(
                                label="⬇️ Download Semua File LPM (Bentuk ZIP)",
                                data=zip_buffer,
                                file_name="Rekap_LPM_Otomatis.zip",
                                mime="application/zip"
                            )
            except Exception as e:
                st.error(f"❌ Terjadi kesalahan saat memproses data LPM: {e}")
