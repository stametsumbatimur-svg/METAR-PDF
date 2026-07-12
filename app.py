import streamlit as st
import pandas as pd
import re
import io
import os
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- DICTIONARY BULAN BAHASA INDONESIA (KAPITAL) ---
BULAN_INDO = {
    1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL",
    5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
    9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER"
}

# --- FUNGSI PARSING DENGAN DETEKSI KOREKSI (COR / CCA / CCB) ---
def parse_metar(sandi_str):
    if pd.isna(sandi_str):
        return None
    sandi_str = sandi_str.replace('\n', ' ').replace('\r', '').strip()
    
    # Cari posisi awal kata METAR
    start_idx = sandi_str.find('METAR')
    if start_idx == -1:
        return None
        
    metar_core = sandi_str[start_idx:].replace('=', '').strip()
    tokens = metar_core.split()
    
    metar, loc, time_str, wind, vis, wx, cloud, t_dp, qnh, rmk = "METAR", "NIL", "NIL", "NIL", "NIL", "NIL", "NIL", "NIL", "NIL", "NOSIG"
    
    # Flag indikator koreksi data
    is_cor = False
    cc_type = ""
    
    remaining_tokens = []
    
    # Fase 1: Identifikasi struktur utama dan flag koreksi
    for t in tokens:
        if t == 'METAR':
            continue
        elif t == 'COR':
            is_cor = True
            continue
        elif re.match(r'^CC[A-Z]$', t):
            cc_type = t.upper()
            continue
        elif re.match(r'^[A-Z]{4}$', t) and loc == "NIL":
            loc = t
            continue
        elif re.match(r'^\d{6}Z$', t) and time_str == "NIL":
            time_str = t
            continue
        else:
            remaining_tokens.append(t)
            
    # Fase 2: Ekstraksi elemen cuaca penerbangan
    if 'CAVOK' in remaining_tokens:
        vis = 'CAVOK'
        wx = ''      
        cloud = ''   
        for t in remaining_tokens:
            if re.match(r'^\d{5}(G\d{2})?KT$', t) or re.match(r'^VRB\d{2}KT$', t) or t == '00000KT':
                wind = t
            elif re.match(r'^\d{2}/\d{2}$', t) or re.match(r'^M\d{2}/\d{2}$', t):
                t_dp = t
            elif re.match(r'^Q\d{4}$', t):
                qnh = t
            elif t in ['NOSIG', 'TEMPO', 'BECMG']:
                rmk = t
    else:
        cloud_list = []
        for t in remaining_tokens:
            if re.match(r'^\d{5}(G\d{2})?KT$', t) or re.match(r'^VRB\d{2}KT$', t) or t == '00000KT':
                wind = t
            elif re.match(r'^\d{4}$', t):
                vis = t
            elif t in ['RA', 'DZ', 'SHRA', 'TSRA', 'TS', 'BR', 'HZ', 'FG', '-RA', '+RA']:
                wx = t
            elif re.match(r'^(FEW|SCT|BKN|OVC)\d{3}$', t) or t in ['NSC', 'SKC', 'CLR']:
                cloud_list.append(t)
            elif re.match(r'^\d{2}/\d{2}$', t) or re.match(r'^M\d{2}/\d{2}$', t):
                t_dp = t
            elif re.match(r'^Q\d{4}$', t):
                qnh = t
            elif t in ['NOSIG', 'TEMPO', 'BECMG']:
                rmk = t
        if cloud_list:
            cloud = " ".join(cloud_list)
            
    return [metar, loc, time_str, wind, vis, wx, cloud, t_dp, qnh, rmk, is_cor, cc_type]

# --- FUNGSI HITUNG BOBOT PRIORITAS DATA ---
def calculate_priority(row):
    score = 0
    if row['is_cor']:
        score = 1
    if row['cc_type'] and len(row['cc_type']) == 3:
        char_code = ord(row['cc_type'][2]) - ord('A')
        score = max(score, 2 + char_code)
    return score

# --- FUNGSI GENERATE PDF ---
def generate_pdf_bytes(df_clean, logo_path):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25
    )
    story = []
    
    styles = getSampleStyleSheet()
    header_text_style = ParagraphStyle(
        'HeaderCenterText', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, leading=15, alignment=1
    )
    rekap_style = ParagraphStyle(
        'RekapStyle', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=11, leading=14, alignment=1
    )
    
    nama_stasiun = df_clean['station_name'].iloc[0].upper() if 'station_name' in df_clean.columns else "STASIUN METEOROLOGI"
    grouped = df_clean.groupby('date_group')
    
    for count, (date, group) in enumerate(grouped):
        if count > 0:
            story.append(PageBreak())
            
        nama_bulan = BULAN_INDO[date.month]
        tanggal_format = f"{date.day:02d} {nama_bulan} {date.year}"
        judul_rekap = f"REKAP DATA METAR: {tanggal_format}".upper()
        text_block = [
            Paragraph("<b>BALAI BESAR METEOROLOGI KLIMATOLOGI DAN GEOFISIKA WILAYAH III</b>", header_text_style),
            Paragraph(f"<b>{nama_stasiun}</b>", header_text_style),
            Paragraph("<b>JL. ADI SUCIPTO NO. 3</b>", header_text_style),
            Paragraph(f"<b>{judul_rekap}</b>", header_text_style)
        ]
        
        if logo_path and os.path.exists(logo_path):
            logo_img = Image(logo_path, width=48, height=48)
            header_table = Table([[logo_img, text_block, ""]], colWidths=[55, 452, 55])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (0,0), 'CENTER'),
                ('ALIGN', (1,0), (1,0), 'CENTER'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('LINEBELOW', (0,0), (-1,-1), 1.5, colors.black),
            ]))
        else:
            header_table = Table([[text_block]], colWidths=[562])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('LINEBELOW', (0,0), (-1,-1), 1.5, colors.black),
            ]))
            
        story.append(header_table)
        story.append(Spacer(1, 10))
        
        
        story.append(Paragraph(f"<b>{judul_rekap}</b>", rekap_style))
        story.append(Spacer(1, 8))
        
        headers = ['METAR', 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK']
        table_data = [headers]
        
        base_table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
            ('TOPPADDING', (0, 0), (-1, -1), 2.5),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]
        
        for idx, row in group.iterrows():
            current_row_idx = len(table_data)
            if row['VIS'] == 'CAVOK':
                row_data = [
                    str(row['METAR']), str(row['LOC']), str(row['TIME']), str(row['WIND']),
                    'CAVOK', '', '', 
                    str(row['T/DP']), str(row['QNH']), str(row['RMK'])
                ]
                base_table_style.append(('SPAN', (4, current_row_idx), (6, current_row_idx)))
            else:
                row_data = [str(row[h]) for h in headers]
            table_data.append(row_data)
            
        col_widths = [45, 40, 55, 75, 42, 40, 105, 45, 50, 65]
        metar_table = Table(table_data, colWidths=col_widths)
        metar_table.setStyle(TableStyle(base_table_style))
        story.append(metar_table)

    doc.build(story)
    buffer.seek(0)
    return buffer

# --- FUNGSI GENERATE EXCEL FORMATTED ---
def generate_excel_bytes(df_clean):
    buffer = io.BytesIO()
    
    headers = ['METAR', 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK', 'datetime']
    df_export = df_clean[headers].copy()
    df_export['datetime'] = df_export['datetime'].dt.strftime('%Y-%m-%d %H:%M:%S')
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Rekap METAR', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Rekap METAR']
        
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(name='Segoe UI', size=10)
                
                col_header = worksheet.cell(row=1, column=cell.column).value
                if col_header in ['METAR', 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'T/DP', 'QNH']:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
        
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    buffer.seek(0)
    return buffer

# --- ANTARMUKA WEB STREAMLIT ---
st.set_page_config(page_title="METAR Data Generator", layout="centered")

st.title("✈️ METAR to PDF & Excel Converter")

LOGO_FILE = "logo_bmkg.png"

if not os.path.exists(LOGO_FILE):
    st.warning(f"⚠️ File gambar '{LOGO_FILE}' tidak terdeteksi di folder utama. Harap pastikan file logo sudah di-upload.")

uploaded_file = st.file_uploader("Upload file CSV", type=["csv"])

if uploaded_file is not None:
    try:
        df = pd.read_csv(uploaded_file)
        
        if 'sandi' not in df.columns or 'data_timestamp' not in df.columns:
            st.error("Format CSV tidak sesuai! Pastikan terdapat kolom 'sandi' and 'data_timestamp'.")
        else:
            with st.spinner("Sedang melakukan validasi data berdasarkan hierarki pembaruan resmi..."):
                parsed_rows = []
                for idx, row in df.iterrows():
                    res = parse_metar(row['sandi'])
                    if res:
                        station = row['station_name'] if 'station_name' in df.columns else "STASIUN METEOROLOGI"
                        msg_id = row['id'] if 'id' in df.columns else idx
                        parsed_rows.append(res + [row['data_timestamp'], station, msg_id])
                        
                columns = ['METAR', 'LOC', 'TIME', 'WIND', 'VIS', 'WX', 'CLOUD', 'T/DP', 'QNH', 'RMK', 'is_cor', 'cc_type', 'raw_timestamp', 'station_name', 'msg_id']
                df_clean = pd.DataFrame(parsed_rows, columns=columns)
                
                df_clean['raw_timestamp'] = df_clean['raw_timestamp'].str.replace(" +0000 UTC", "", regex=False)
                df_clean['datetime'] = pd.to_datetime(df_clean['raw_timestamp'])
                
                df_clean = df_clean[df_clean['datetime'].dt.minute == 0]
                df_clean['priority_score'] = df_clean.apply(calculate_priority, axis=1)
                df_clean = df_clean.sort_values(by=['datetime', 'priority_score', 'msg_id'])
                df_clean = df_clean.drop_duplicates(subset=['datetime'], keep='last')
                
                df_clean['date_group'] = df_clean['datetime'].dt.date
                df_clean = df_clean.sort_values(by='datetime').reset_index(drop=True)
                
                if df_clean.empty:
                    st.warning("Tidak ditemukan data dengan menit :00 (per jam) di dalam file ini.")
                else:
                    st.success(f"Berhasil!")
                    
                    st.subheader("Preview Data Tervalidasi")
                    # AMAN: Menghapus parameter width agar mengikuti auto-layout bawaan Streamlit
                    st.dataframe(df_clean[['METAR', 'LOC', 'TIME', 'WIND', 'VIS', 'CLOUD', 'T/DP', 'QNH']].head(10))
                    
                    pdf_data = generate_pdf_bytes(df_clean, LOGO_FILE)
                    excel_data = generate_excel_bytes(df_clean)
                    
                    first_date = df_clean['date_group'].iloc[0]
                    nama_file_base = f"REKAP_METAR_{first_date.day:02d}_{BULAN_INDO[first_date.month]}_{first_date.year}"
                    
                    st.write("---")
                    st.subheader("Unduh Laporan Ekspor")
                    
                    col_pdf, col_xlsx = st.columns(2)
                    
                    with col_pdf:
                        # AMAN: Menghapus parameter ukuran yang memicu Segfault biner
                        st.download_button(
                            label="📥 Download PDF",
                            data=pdf_data,
                            file_name=f"{nama_file_base}.pdf",
                            mime="application/pdf"
                        )
                        
                    with col_xlsx:
                        # AMAN: Menghapus parameter ukuran yang memicu Segfault biner
                        st.download_button(
                            label="📊 Download Excel Spreadsheet",
                            data=excel_data,
                            file_name=f"{nama_file_base}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
    except Exception as e:
        st.error(f"Terjadi kesalahan: {e}")
